#!/usr/bin/env python
# coding: utf-8

# In[ ]:


def run_process(a1, a2, a3, b1, b2, b3, email):
    from selenium.common.exceptions import ElementClickInterceptedException
    from urllib3.util.timeout import Timeout
    #PASSWORDS
    surya_link='https://client.suryacargo.com/'
    surya_username="AMAZON"
    surya_pass="AMZS001"
    indigo_link='https://6ecargo.goindigo.in/Login.aspx'
    indigo_username="AMAZCNTRAL"
    indigo_pass="Cargo@2024"
    spice_link='https://spicexpress.kargo360tech.com/'
    spice_username="ASACENDO"
    spice_pass="Kargo@12345"
    pnq_quip='https://quip-amazon.com/4b5aAZJQiubo/QJ-Dedicated-Charter-Internal-MIS-Feb02-Mar02#temp:C:FAIe067c2e081cf4922b6f510489'
    proverne='https://qo.proverne.com/cargo/login'
    prov_email='amazon-cargo@amazon.com'
    prov_key="Cargoin@2023"
    #PNQ_vol
    from selenium import webdriver
    from webdriver_manager.firefox import GeckoDriverManager
    from selenium.webdriver.common.service import Service
    from selenium.webdriver.firefox.service import Service as GeckoDriverService
    import os
    directory_1 = r'C:\Project\CLH_Data'
    if not os.path.exists(directory_1):
        os.makedirs(directory_1)
        print("Directory created successfully.")
    else:
        print("Directory already exists.")
    from webdriver_manager.firefox import GeckoDriverManager
    GECKODRIVER_PATH=gecko_driver_path = GeckoDriverManager().install()
    FIREFOX_PATH=firefox_binary_path = r'C:\Program Files\Mozilla Firefox\firefox.exe'
    download_dir = r'C:\Downloads'
    destination_dir = r'C:\Project\CLH_Data'
    destination_dir_1= r'C:\Project\CLH_Data\charter'
    clh_file_path = r'C:\Project\CLH_Data\CLH-input.xlsx'
    flight_data_file_path = r'C:\Project\CLH_Data\Flight_data.xlsx'
    allied_df_path=r'C:\Project\CLH_Data\Allied.xlsx'
    indigo_file_path = r'C:\Project\CLH_Data\Indigo.xlsx'
    surya_file=r'C:\Project\CLH_Data\Surya.xlsx'
    spicejet_file=r'C:\Project\CLH_Data\Spicejet.xlsx'
    pobc_file=r'C:\Project\CLH_Data\pobc.xlsx'
    krbl_file=r'C:\Project\CLH_Data\krbl.xlsx'
    krbla_file=r'C:\Project\CLH_Data\krbl-airasia.xlsx'
    index_file=r'C:\Project\CLH_Data\Index.xlsx'
    labriynth_file=r'C:\Project\CLH_Data\RL Sea movement .xlsx'
    labriynth_1_file=r'C:\Project\CLH_Data\Sea Movement .xlsx'
    rates_file=r'C:\Project\CLH_Data\rates.xlsx'
    output_file = r'C:\Project\CLH_Data\output_belly.xlsx'
    pnq_vols=file_path_1=r'C:\Project\CLH_Data\pnq_vols.xlsx'
    quip_file=r'C:\Project\CLH_Data\quip.xlsx'
    qj_mis_file=r'C:\Project\CLH_Data\qj_mis.xlsx'
    otp_report_file=r'C:\Project\CLH_Data\otp_report_1.xlsm'
    otp_report=file_path_2=r'C:\Project\CLH_Data\otp_report.xlsx'
    output_char= r'C:\Project\CLH_Data\output_charter.xlsx'
    import getpass
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    import time
    from selenium.webdriver.firefox.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import os
    firefox_options = webdriver.FirefoxOptions()
    firefox_options.headless = True  # Run in headless mode
    timeout=30
    firefox_options.binary_location = firefox_binary_path
    capabilities = {
        "browserName": "firefox",
        "moz:firefoxOptions": {
            "args": [],
            "log": {"level": "trace"},
            "prefs": {
                "network.proxy.http": "localhost",
                "network.proxy.http_port": "8080",
                "network.proxy.type": 1
            }
        }
    }
    from selenium import webdriver
    firefox_options.set_preference("browser.download.folderList", 2)  # 0: Desktop, 1: Downloads, 2: Custom
    firefox_options.set_preference("browser.download.manager.showWhenStarting", False)
    firefox_options.set_preference("browser.download.dir", "C:\\Project\\CLH_Data")  # Specify your desired download directory
    firefox_options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    service = GeckoDriverService(executable_path=gecko_driver_path)
    driver = webdriver.Firefox(service=service, options=firefox_options)
    max_attempts = 4
    driver.get(pnq_quip)
    time.sleep(5)
    email_field = driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/div/input')
    email_field.send_keys(email)
    for attempt in range(1, max_attempts + 1):
        try:
            submit_button = driver.find_element(By.XPATH, '//*[@id="email-submit"]')
            submit_button.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    time.sleep(2)
    for attempt in range(1, max_attempts + 1):
        try:
            button1 = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div/div[2]/div[2]/div[1]/div[1]/div[1]/div[2]/div/button[1]/div')
            button1.click()
            break  # Exit the loop if the click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise the exception if max attempts reached
            time.sleep(3)  
    
    for attempt in range(1, max_attempts + 1):
        try:
            button2 = driver.find_element(By.XPATH, '/html/body/div[6]/div/div/div/div[11]')
            button2.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    time.sleep(2)
    for attempt in range(1, max_attempts + 1):
        try:
            final_button = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div/div[4]/div[2]')
            final_button.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    time.sleep(18)
    driver.quit()
    files_found = [f for f in os.listdir(destination_dir) if f.startswith("QJ-Dedicated-Charter-Internal")]

    # Rename files to "pnq_vols"
    for file in files_found:
        old_file_path = os.path.join(destination_dir, file)
        new_file_path = os.path.join(destination_dir, "quip.xlsx")  # New name
        os.rename(old_file_path, new_file_path)
        print(f"Renamed: {file} to quip.xlsx")
    import os
    import shutil

    #Surya Cargo
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import time
    import os
    import shutil
    import time
    import urllib.request
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.firefox.options import Options
    a = str(a1) + "-" + str(a2) + "-" + str(a3) 
    b = str(b1) + "-" + str(b2) + "-" + str(b3)
    firefox_options = webdriver.FirefoxOptions()
    firefox_options.headless = True  # Run in headless mode
    #firefox_options.add_argument('--no-sandbox')  # Bypass OS security model
    firefox_options.add_argument('--disable-dev-shm-usage')
    timeout=30

    firefox_options = Options()
    firefox_options.binary_location = firefox_binary_path 
    capabilities = webdriver.DesiredCapabilities().FIREFOX.copy()
    capabilities['pageLoadStrategy'] = 'eager'
    firefox_options.set_preference("browser.download.folderList", 2)  # 0: Desktop, 1: Downloads, 2: Custom
    firefox_options.set_preference("browser.download.manager.showWhenStarting", False)
    firefox_options.set_preference("browser.download.dir", "C:\\Project\\CLH_Data")  # Specify your desired download directory
    firefox_options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    service = GeckoDriverService(executable_path=gecko_driver_path)
    driver = webdriver.Firefox(service=service, options=firefox_options)
    
    driver.set_page_load_timeout(60)
    driver.get(surya_link)
    time.sleep(2)
    driver.maximize_window()
    username_field = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='UserName']"))
    )
    username_field.send_keys(surya_username)
    password_field = driver.find_element(By.XPATH, "//*[@id='Password']")
    password_field.send_keys(surya_pass)
    for attempt in range(1, max_attempts + 1):
        try:
            signin_button = driver.find_element(By.XPATH, "//button[@class='SignInKls']")
            signin_button.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    time.sleep(3)
    mis_tab = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'MIS')]")))
    driver.execute_script("arguments[0].scrollIntoView(true);", mis_tab)
    mis_tab.click()
    time.sleep(3)
    customer_mis_option = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/div[5]/div[1]/div[1]/div/div/nav[1]/ul/li[1]/div/a[2]")))
    driver.execute_script("arguments[0].scrollIntoView(true);", customer_mis_option)
    customer_mis_option.click()
    time.sleep(2)
    for attempt in range(1, max_attempts + 1):
        try:
            date_field = driver.find_element(By.XPATH, '//*[@id="FromDateFRR"]')
            date_field.clear()
            break  # Exit the loop if clear is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)  
    date_field.send_keys(a) 
    for attempt in range(1, max_attempts + 1):
        try:
            date_field_1 = driver.find_element(By.XPATH, '//*[@id="ToDateFRR"]')
            date_field_1.clear()
            break  # Exit the loop if clear is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)  
    date_field_1.send_keys(b) 
    for attempt in range(1, max_attempts + 1):
        try:
            run_report_button = driver.find_element(By.XPATH, '//*[@id="CustMisRemarktbl"]/tbody/tr[2]/td[9]/input')
            run_report_button.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    time.sleep(3)
    for attempt in range(1, max_attempts + 1):
        try:
            export_button = driver.find_element(By.XPATH, '//*[@id="GridSuryaCustomerMISReport"]/div[1]/a[1]')
            export_button.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    time.sleep(5)
    driver.quit()  
    files_found = [f for f in os.listdir(destination_dir) if f.startswith("Customer MIS")]

    # Rename files to "pnq_vols"
    for file in files_found:
        old_file_path = os.path.join(destination_dir, file)
        new_file_path = os.path.join(destination_dir, "Surya.xlsx")  # New name
        os.rename(old_file_path, new_file_path)
        print(f"Renamed: {file} to Surya")

    import pandas as pd
    df = pd.read_excel(quip_file,sheet_name='MIS')
    df['Flight Date'] = pd.to_datetime(df['Flight Date'], errors='coerce', format='%m/%d/%Y')
    df = df.dropna(subset=['Flight Date'])
    b = str(b1) + "/" + str(b2) + "/" + str(b3)
    df = df[(df['Flight Date'] >= a) & (df['Flight Date'] <= b)]
    df = df[df['Destination'] == 'PNQ']
    df.to_excel('pnq_vols.xlsx', index=False)
    df = df[['AWB#', 'Lane']]
    df = df.rename(columns={'AWB#': 'Awb No', 'Lane': 'lane_updated'})
    df.head(5)
    df.to_excel(pnq_vols, index=False)
    #print(f"Filtered data saved to pnq_vols.xlsx. Date range: {a.strftime('%m/%d/%Y')} to {b.strftime('%m/%d/%Y')}")
    print(a)
    print(b)

    #INDIGO
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import time
    import os
    import shutil
    from datetime import datetime

    firefox_options = webdriver.FirefoxOptions()
    firefox_options.headless = True  # Run in headless mode
    timeout=30
    firefox_options.binary_location = FIREFOX_PATH
    firefox_options.set_capability("acceptInsecureCerts", True)
    firefox_options.set_preference("browser.download.folderList", 2)  # 0: Desktop, 1: Downloads, 2: Custom
    firefox_options.set_preference("browser.download.manager.showWhenStarting", False)
    firefox_options.set_preference("browser.download.dir", "C:\\Project\\CLH_Data")  # Specify your desired download directory
    firefox_options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    service = GeckoDriverService(executable_path=gecko_driver_path)
    driver = webdriver.Firefox(service=service, options=firefox_options)
    a1 = int(a1)
    a2 = int(a2)
    a3 = int(a3)
    b1 = int(b1)
    b2 = int(b2)
    b3 = int(b3)
    from_date = datetime(a3, a2, a1)
    to_date = datetime(b3, b2, b1)
    date_str = from_date.strftime('%d/%m/%Y')
    date_str_1 = to_date.strftime('%d/%m/%Y')
    #a = str(a1) + "/" + str(a2) + "/" + str(a3)
    #b = str(b1) + "/" + str(b2) + "/" + str(b3)

    driver.get(indigo_link)
    time.sleep(2)
    driver.maximize_window()
    username_field = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtUserName"]'))
    )
    username_field.send_keys(indigo_username)
    password_field = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtPwd"]')
    password_field.send_keys(indigo_pass)
    for attempt in range(1, max_attempts + 1):
        try:
            signin_button = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnLogin"]')
            signin_button.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    Finance_tab = WebDriverWait(driver, 8).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="menu"]/ul/li[3]/a/span')))
    driver.execute_script("arguments[0].scrollIntoView(true);", Finance_tab)
    Finance_tab.click()
    Rate_Audit = WebDriverWait(driver, 8).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="menu"]/ul/li[3]/ul/li[1]/a/span')))
    driver.execute_script("arguments[0].scrollIntoView(true);", Rate_Audit)
    Rate_Audit.click()
    for attempt in range(1, max_attempts + 1):
        try:
            date_field = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtbillingfrom1_txtDate"]')
            date_field.clear()
            break  # Exit the loop if clear is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)  
    date_field.send_keys(date_str) 
    for attempt in range(1, max_attempts + 1):
        try:
            date_field_1 = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtbillingto1_txtDate"]')
            date_field_1.clear()
            break  # Exit the loop if clear is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)  
    date_field_1.send_keys(date_str_1) 
    checkbox = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_chkShipment"]')
    if checkbox.is_selected():
        checkbox.click()
    List= WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnSearch"]')))
    driver.execute_script("arguments[0].scrollIntoView(true);", List)
    List.click() 
    time.sleep(35)
    #WebDriverWait(driver, 10).until_not(EC.visibility_of_element_located((By.ID, "msgfade")))
    #export_button = WebDriverWait(driver, 65).until(
    #    EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnExportBilling"]')))
    #export_button.click()
    for attempt in range(1, max_attempts + 1):
        try:
            export_button = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnExportBilling"]')
            export_button.click()
            break  # Exit the loop if clear is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(6)
    time.sleep(70)
    files_found = [f for f in os.listdir(destination_dir) if f.startswith("BillingExport")]

    # Rename files to "pnq_vols"
    for file in files_found:
        old_file_path = os.path.join(destination_dir, file)
        new_file_path = os.path.join(destination_dir, "Indigo.xlsx")  # New name
        os.rename(old_file_path, new_file_path)
        print(f"Renamed: {file} to Indigo")
    time.sleep(10)
    driver.quit()

    #Spicejet
    from datetime import datetime
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import time
    import os
    import shutil
    import time
    import urllib.request
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.firefox.options import Options
    from selenium.webdriver.support.ui import Select
    from selenium.webdriver.common.action_chains import ActionChains

    a1 = int(a1)
    a2 = int(a2)
    a3 = int(a3)
    b1 = int(b1)
    b2 = int(b2)
    b3 = int(b3)
    from_date = datetime(a3, a2, a1)
    to_date = datetime(b3, b2, b1)
    date_str = from_date.strftime('%d-%m-%Y')
    date_str_1 = to_date.strftime('%d-%m-%Y')


    firefox_options = webdriver.FirefoxOptions()
    firefox_options.binary_location = FIREFOX_PATH
    firefox_options.set_capability("moz:firefoxOptions", {"binary": FIREFOX_PATH})
    firefox_options.set_preference("browser.download.folderList", 2)  # 0: Desktop, 1: Downloads, 2: Custom
    firefox_options.set_preference("browser.download.manager.showWhenStarting", False)
    firefox_options.set_preference("browser.download.dir", "C:\\Project\\CLH_Data")  # Specify your desired download directory
    firefox_options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    service = GeckoDriverService(executable_path=gecko_driver_path)
    driver = webdriver.Firefox(service=service, options=firefox_options)
    time.sleep(5)
    driver.get(spice_link)
    
    time.sleep(2)
    driver.maximize_window()
    username_field = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '/html/body/app-root/app-login/div/div/div/div/div[2]/div[2]/div[2]/form/div/div/div[2]/input'))
    )
    username_field.send_keys(spice_username)
    password_field = driver.find_element(By.XPATH, '/html/body/app-root/app-login/div/div/div/div/div[2]/div[2]/div[2]/form/div/div/div[3]/input')
    password_field.send_keys(spice_pass)
    for attempt in range(1, max_attempts + 1):
        try:
            signin_button = driver.find_element(By.XPATH, '/html/body/app-root/app-login/div/div/div/div/div[2]/div[2]/div[2]/form/div/div/div[4]/button')
            signin_button.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    time.sleep(2)

    Reports_tab = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/app-top-navigation/nav/div[2]/ul[3]/li/a/span[1]')))
    driver.execute_script("arguments[0].scrollIntoView(true);", Reports_tab)
    Reports_tab.click()
    time.sleep(3)
    Standard_tab = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/app-top-navigation/nav/div[2]/ul[3]/li/ul/li[2]/a')))
    driver.execute_script("arguments[0].scrollIntoView(true);", Standard_tab)
    Standard_tab.click()
    time.sleep(3)
    AWB_ton = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/app-top-navigation/nav/div[2]/ul[3]/li/ul/li[2]/ul/li/a')))
    driver.execute_script("arguments[0].scrollIntoView(true);", AWB_ton)
    AWB_ton.click()

    time.sleep(3)
    driver.execute_script("window.scrollTo(0, 0)")
    time.sleep(3)
    for attempt in range(1, max_attempts + 1):
        try:
            date_field = driver.find_element(By.XPATH, '/html/body/app-root/app-tonnage/div/div/div/div[2]/div/div/div/div/form/div/div/div/div/div[4]/div/input')
            date_field.clear()
            break  # Exit the loop if clear is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3) 
    date_field.send_keys(date_str)
    time.sleep(3)
    for attempt in range(1, max_attempts + 1):
        try:
            date_field_1 = driver.find_element(By.XPATH, '/html/body/app-root/app-tonnage/div/div/div/div[2]/div/div/div/div/form/div/div/div/div/div[5]/div/input')
            date_field_1.clear()
            break  # Exit the loop if clear is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)  
    date_field_1.send_keys(date_str_1) 
    body = driver.find_element(By.XPATH, '//body')
    ActionChains(driver).move_to_element(body).click().perform()
    time.sleep(5)

    driver.execute_script("window.scrollTo(0, 0)")
    download = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/app-tonnage/div/div/div/div[2]/div/div/div/div/form/div/div/div/div/div[6]/div/label/button/i')))
    driver.execute_script("arguments[0].scrollIntoView(true);", download)
    driver.execute_script("window.scrollTo(0, 0)")
    download.click()
    time.sleep(6)
    files_found = [f for f in os.listdir(destination_dir) if f.startswith("tonnage-report")]
    time.sleep(5)
    # Rename files to "Spicejet"
    for file in files_found:
        old_file_path = os.path.join(destination_dir, file)
        new_file_path = os.path.join(destination_dir, "Spicejet.xlsx")  # New name
        os.rename(old_file_path, new_file_path)
        print(f"Renamed: {file} to Spicejet")
    time.sleep(10)
    
    driver.quit()
    
    #POBC
    import os
    import win32com.client
    import pythoncom
    try:
        pythoncom.CoInitialize()
    except pythoncom.com_error:
        pass
    def download_latest_email_attachment(subject_prefix, body_keyword, destination_dir):
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6)
        items = inbox.Items
        items.Sort("[ReceivedTime]", True)
        for mail in items:
            if mail.Subject.startswith(subject_prefix):
                if body_keyword.lower() in mail.Body.lower():
                    if mail.Attachments.Count > 0:
                        for attachment in mail.Attachments:
                            if attachment.FileName.endswith('.xlsx'):
                                attachment.SaveAsFile(os.path.join(destination_dir, "pobc.xlsx"))
                                print(f"Excel attachment saved successfully from email with subject: '{mail.Subject}'.")
                                return
                    else:
                        print("No attachment found in the latest email.")
                        return
        print("No email found with the specified subject prefix and keyword in the body.")
    subject_prefix = "Amazon Status - India (CLH )"
    body_keyword = "Patel Integrated Logistics"


    download_latest_email_attachment(subject_prefix, body_keyword, destination_dir)

    #KRBL-Airasia
    import os
    import win32com.client

    def download_latest_email_attachment(subject_prefix, body_keyword,destination_dir):
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6) 
        items = inbox.Items
        items.Sort("[ReceivedTime]", True)
        for mail in items:
            if mail.Subject.startswith(subject_prefix):
                if body_keyword.lower() in mail.Body.lower():
                    if mail.Attachments.Count > 0:
                        for attachment in mail.Attachments:
                            if attachment.FileName.endswith('.xlsx'):
                                attachment.SaveAsFile(os.path.join(destination_dir, "krbl-airasia.xlsx"))
                                print(f"Excel attachment saved successfully from email with subject: '{mail.Subject}'.")
                                return
                    else:
                        print("No attachment found in the latest email.")
                        return
        print("No email found with the specified subject prefix and keyword in the body.")

    subject_prefix = "AirAsia Amazon Pace MIS"
    body_keyword = "Kreative"

    download_latest_email_attachment(subject_prefix, body_keyword, destination_dir)


    #KRBL
    import os
    import win32com.client

    def download_latest_email_attachment(subject_prefix, body_keyword, destination_dir):
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6)
        items = inbox.Items
        items.Sort("[ReceivedTime]", True)
        for mail in items:
            if mail.Subject.startswith(subject_prefix):
                if body_keyword.lower() in mail.Body.lower():
                    if mail.Attachments.Count > 0:
                        for attachment in mail.Attachments:
                            if attachment.FileName.endswith('.xlsx'):
                                attachment.SaveAsFile(os.path.join(destination_dir, "krbl.xlsx"))
                                print(f"Excel attachment saved successfully from email with subject: '{mail.Subject}'.")
                                return
                    else:
                        print("No attachment found in the latest email.")
                        return
        print("No email found with the specified subject prefix and keyword in the body.")

    subject_prefix = "Amazon MIS as on"
    body_keyword = "Kreative"


    download_latest_email_attachment(subject_prefix, body_keyword, destination_dir)

    #INDEX
    import os
    import win32com.client

    def download_latest_email_attachment(subject_prefix, body_keyword, destination_dir):
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6) 
        items = inbox.Items
        items.Sort("[ReceivedTime]", True) 

        for mail in items:
            if mail.Subject == subject_prefix:
                if body_keyword.lower() in mail.Body.lower():
                    if mail.Attachments.Count > 0:
                        for attachment in mail.Attachments:
                            if attachment.FileName.endswith('.xlsx'):
                                attachment.SaveAsFile(os.path.join(destination_dir, "Index.xlsx"))
                                print(f"Excel attachment saved successfully from email with subject: '{mail.Subject}'.")
                                return
                    else:
                        print("No Excel attachment found in the latest email.")
                        return
        print("No email found with the specified subject prefix and keyword in the body.")


    subject_prefix = "[EXTERNAL] [UNVERIFIED SENDER] BOOKING/UPLIFTING // AMAZON // DATED"
    body_keyword= "IndEx"

    download_latest_email_attachment(subject_prefix,body_keyword, destination_dir)


    #Allied
    import os
    import win32com.client

    def download_latest_email_attachment(subject_prefix, body_keyword, destination_dir):
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6) 
        items = inbox.Items
        items.Sort("[ReceivedTime]", True)
        for mail in items:
            if mail.Subject.startswith(subject_prefix):
                if body_keyword.lower() in mail.Body.lower():
                    if mail.Attachments.Count > 0:
                        for attachment in mail.Attachments:
                            if attachment.FileName.endswith('.xlsx'):
                                attachment.SaveAsFile(os.path.join(destination_dir, "Allied.xlsx"))
                                print(f"Excel attachment saved successfully from email with subject: '{mail.Subject}'.")
                                return
                    else:
                        print("No attachment found in the latest email.")
                        return
        print("No email found with the specified subject prefix and keyword in the body.")

    subject_prefix = "Amazon MIS Data for"
    body_keyword = "Unit 34, Adarsh Industrial Estate, Chakala, Andheri (E), Mumbai – 400099, India"


    download_latest_email_attachment(subject_prefix, body_keyword, destination_dir)

    import os
    import win32com.client
    import logging

    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    def scrape_emails(subject_prefix, body_keyword, destination_dir, outlook_folder=6):
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        folder = outlook.GetDefaultFolder(outlook_folder)
        items = folder.Items
        items.Sort("[ReceivedTime]", True)

        for mail in items:
            if mail.Subject.lower().startswith(subject_prefix.lower()) and body_keyword.lower() in mail.Body.lower():
                if mail.Attachments.Count > 0:
                    for attachment in mail.Attachments:
                        attachment_name_parts = attachment.FileName.split('.')
                        attachment_name = attachment_name_parts[0].split('-')[0] + '.' + attachment_name_parts[-1]
                        try:
                            attachment.SaveAsFile(os.path.join(destination_dir, attachment_name))
                            logging.info("Attachment saved successfully: %s", attachment.FileName)
                        except Exception as e:
                            logging.error("Failed to save attachment: %s", e)
                else:
                    logging.info("No attachment found in the latest email.")
                return
        logging.info("No email found with the specified subject prefix and keyword in the body.")

    # Define subject prefix, body keyword, and destination directory
    subject_prefix = "RE: SEA_ATS_STD_Report"
    body_keyword = "SEA ATS"


    # Call the function to scrape emails
    scrape_emails(subject_prefix, body_keyword, destination_dir)


    #Charter-qj_mis
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains
    import urllib.request
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.firefox.service import Service
    from selenium.webdriver.firefox.options import Options
    from selenium.webdriver.support.ui import Select
    import time 


    firefox_options = webdriver.FirefoxOptions()
    firefox_options.binary_location = FIREFOX_PATH
    firefox_options.set_capability("moz:firefoxOptions", {"binary": FIREFOX_PATH})
    firefox_options.set_capability("acceptInsecureCerts", True)
    firefox_options.set_preference("browser.download.folderList", 2)  # 0: Desktop, 1: Downloads, 2: Custom
    firefox_options.set_preference("browser.download.manager.showWhenStarting", False)
    firefox_options.set_preference("browser.download.dir", "C:\\Project\\CLH_Data")  # Specify your desired download directory
    firefox_options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    service = GeckoDriverService(executable_path=gecko_driver_path)
    driver = webdriver.Firefox(service=service, options=firefox_options)

    driver.get(proverne)
    a = str(a2) + "/" + str(a1) + "/" + str(a3)

    email_input = driver.find_element(By.XPATH, '//*[@id="mat-input-2"]')
    email_input.send_keys(prov_email)
    password_input = driver.find_element(By.XPATH, "//*[@id=\"mat-input-3\"]")
    password_input.send_keys(prov_key)
    for attempt in range(1, max_attempts + 1):
        try:
            login_button = driver.find_element(By.XPATH, '//*[@id="login-form"]/form/div/button')
            login_button.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    time.sleep(2)
    for attempt in range(1, max_attempts + 1):
        try:
            reports_button = driver.find_element(By.XPATH, '//*[@id="container-3"]/toolbar/mat-toolbar/mat-toolbar-row[2]/div/div[5]/a/span/span')
            reports_button.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    for attempt in range(1, max_attempts + 1):
        try:
            report_link = driver.find_element(By.XPATH, '//*[@id="mat-menu-panel-34"]/div/button[1]/span')
            report_link.click()
            break  # Exit the loop if click is successful
        except (NoSuchElementException, ElementClickInterceptedException):
            if attempt == max_attempts:
                raise  # Raise exception if max_attempts reached
            time.sleep(3)
    run_report_button = driver.find_element(By.XPATH, '//*[@id="mat-menu-panel-35"]/div/button[6]/span')
    run_report_button.click()
    time.sleep(2)
    search_templates_button = driver.find_element(By.XPATH, '//*[@id="cdk-accordion-child-0"]/div/div/div[3]/app-submitbutton/div/button/span')
    search_templates_button.click()
    time.sleep(5)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(5)
    awb_flight_reports_link = driver.find_element(By.XPATH, '//*[@id="cdk-accordion-child-1"]/div/mat-table/mat-row[5]/mat-cell[2]')
    awb_flight_reports_link.click()
    time.sleep(2)
    Attribute=driver.find_element(By.XPATH, '//*[@id="cdk-accordion-child-2"]/div/div/div[2]/form/div/mat-table/mat-row/mat-cell[1]/mat-form-field/div/div[1]')
    Attribute.click()
    button = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="mat-option-205"]/span')))
    driver.execute_script("arguments[0].scrollIntoView(true);", button)
    button.click()
    Attribute_1=driver.find_element(By.XPATH, '//*[@id="mat-select-12"]/div/div[2]/div')
    Attribute_1.click()
    ge_option = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="mat-option-216"]/span')))
    driver.execute_script("arguments[0].scrollIntoView(true);", ge_option)
    ge_option.click()
    From_Date= driver.find_element(By.XPATH,'//*[@id="mat-input-9"]')
    From_Date.click()
    From_Date.send_keys(a)
    generate_time_button = driver.find_element(By.XPATH, '//*[@id="cdk-accordion-child-2"]/div/div/div[3]/div/app-submitbutton/div/button/span')
    generate_time_button.click()
    time.sleep(5)
    try:
        alert = WebDriverWait(driver, 5).until(EC.alert_is_present())
        alert.accept()
        print("Alert accepted.")
    except Exception as e:
        print("No alert present.")
    time.sleep(10)

    export_report_button = driver.find_element(By.XPATH, '/html/body/app/vertical-layout-1/div/div/div/div/content/app-runreporttemplate/div/div/div/div/div/div/form/mat-expansion-panel[3]/div/div/div/div[3]/div/app-submitbutton[1]/div/button/span')
    export_report_button.click()
    time.sleep(15)
    files_found = [f for f in os.listdir(destination_dir) if f.endswith(".xls")]
   # Rename files to "qj_mis.xlsx"
    for file in files_found:
        old_file_path = os.path.join(destination_dir, file)
        new_file_path = os.path.join(destination_dir, "qj_mis.xlsx")  # New name
        os.rename(old_file_path, new_file_path)
        print(f"Renamed: {file} to qj_mis")

    driver.quit()


    import os
    import win32com.client

    def download_latest_email_attachment(subject_prefix, body_keyword, destination_dir):
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6)  
        items = inbox.Items
        items.Sort("[ReceivedTime]", True)
        for mail in items:
            if mail.Subject.startswith(subject_prefix):
                if body_keyword.lower() in mail.Body.lower():
                    if mail.Attachments.Count > 0:
                        for attachment in mail.Attachments:
                            if attachment.FileName.startswith('OTP Report') and attachment.FileName.endswith('.xlsm'):
                                attachment.SaveAsFile(os.path.join(destination_dir, "otp_report_1.xlsm"))
                                print(f"Excel attachment saved successfully from email with subject: '{mail.Subject}'.")
                                return
                    else:
                        print(f"No attachment found in the latest email with subject: '{mail.Subject}'.")
                        return
        print("No email found with the specified subject prefix and keyword in the body.")


    subject_prefix = "OTP, CUT ,TAXIBOT & ULD REPORTS"
    body_keyword = "Quikjet Cargo Airlines Private Limited"


    download_latest_email_attachment(subject_prefix, body_keyword, destination_dir)
    # Data Manipulation

    #qj_mis
    df_1=pd.read_excel(qj_mis_file,sheet_name='Sales')
    df_1 = df_1.drop_duplicates(subset='Awb No')
    df_1.head(5)
    df_1.to_excel(qj_mis_file, sheet_name='Sales', index=False)

    #OTP_report
    import pandas as pd
    df_2=pd.read_excel(otp_report_file,sheet_name='Apr 24',skiprows=4)
    df_2.drop(0, axis=0, inplace=True)
    columns_to_keep = ['SN', 'Day', 'Aircraft', 'Flt.No', 'Sector', 'BT', 'Bowser Uplift Fuel           (LTRS)', 'Total Fuel Burn (Litre)']
    df_2.columns = df_2.columns.str.strip() 
    df_2_filtered = df_2[columns_to_keep]
    df_2_filtered = df_2_filtered[(df_2_filtered['BT'] != 0) & (df_2_filtered['BT'].notna())]
    sectors_to_remove = ['HYD-VGA', 'VGA-STV', 'STV-DEL']
    df_2_filtered = df_2_filtered[~df_2_filtered['Sector'].isin(sectors_to_remove)]
    df_2_filtered = df_2_filtered.dropna(subset=['Sector'])
    df_2_filtered.rename(columns={'Bowser Uplift Fuel           (LTRS)': 'Bowser Uplift Fuel(LTRS)'}, inplace=True)
    df_2_filtered.head()
    df_2_filtered.to_excel(otp_report, index=False)
    destination_dir = r'C:\Project\CLH_Data'
    file_path = os.path.join(destination_dir, 'CLH-inputs.xlsx')

    #Clearing old data sheets
    from openpyxl import load_workbook
    import pandas as pd
    import numpy as np
    from openpyxl.worksheet.worksheet import Worksheet 
      # Check if the file exists before attempting to delete
    
    if os.path.exists(clh_file_path):
        os.remove(clh_file_path)
        print(f"File {clh_file_path} deleted successfully")
    else:
        print(f"File {clh_file_path} does not exist")
        

  

    #Allied
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook


    flight_data = pd.read_excel(flight_data_file_path)
    allied_df= pd.read_excel(allied_df_path,sheet_name='MARCH')
    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh = pd.DataFrame(columns=columns)

    clh['Date'] = allied_df['Flight Date'].dt.strftime('%m/%d/%Y')
    clh['Origin'] = allied_df['Origin']
    clh['Dest'] = allied_df['Dest']
    clh['Lane'] = allied_df['Origin'] + '-' + allied_df['Dest']
    clh['Co-Loader'] = 'Allied Express'
    clh['Airline'] = allied_df['Airlines']
    clh['Flight'] = allied_df['Flt No'].apply(lambda x: x[:2] + '0' + x[2:] if len(x) == 5 else x)
    clh['AWBnum'] = allied_df['Mawb No.'].str.split('-', n=1).str[1]
    clh['CDnum'] = clh['AWBnum']
    clh['Pcs'] = allied_df['Bags']
    clh['Gross weight'] = allied_df['Gross Wght']
    clh['Dim weight'] = allied_df['Dim Wght']
    clh['Chg'] = allied_df['Dim Wght']
    clh['Mode'] = 'Dedicated'
    clh['ETD'] = allied_df['ETD']
    clh['ETA'] = allied_df['ETA']
    clh['Approver'] = '-'
    clh['Mode']='Dedicated'
    #clh = clh.dropna(subset=["Date", "AWBnum"], how="all")
    clh = clh.dropna(subset=['AWBnum']).drop(clh[clh['AWBnum'].isin(['', '-'])].index).reset_index(drop=True)

    #print(clh['Flight'])
    with pd.ExcelWriter(clh_file_path, mode='w', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Allied')


    #Indigo
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook


    indigo = pd.read_excel(indigo_file_path)
    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh = pd.DataFrame(columns=columns)

    indigo['AWBDate'] = pd.to_datetime(indigo['AWBDate'], format='%d/%m/%Y')
    clh['Date'] = pd.to_datetime(indigo['AWBDate'])
    clh['Date'] = pd.to_datetime(clh['Date'], format='%d/%m/%Y').dt.strftime('%d/%m/%Y')
    clh['Date'] = pd.to_datetime(clh['Date'], format='%d/%m/%Y').dt.strftime('%m/%d/%Y')

    clh['Origin'] = indigo['Origin']
    clh['Dest'] = indigo['Destination']
    clh['Lane'] = indigo['Origin'] + '-' + indigo['Destination']
    clh['Co-Loader'] = 'IndiGo Air'
    clh['Airline'] = 'IndiGo' 
    clh['Flight']= indigo['FltNo'].apply(lambda x: x[:2] + '0' + x[2:] if len(x) == 5 else x)
    clh['AWBnum'] = indigo['AWBNumber']  
    clh['CDnum'] = indigo['AWBNumber']  
    clh['Pcs'] = indigo['Pieces']  
    clh['Gross weight'] = indigo['Gross Weight']  
    clh['Dim weight'] = indigo['ChargedWeight']  
    clh['Chg'] = indigo['ChargedWeight']  
    clh['Mode'] = np.where(indigo['CommodityCode'] == 'XPS', 'Express', 'Dedicated')
    clh['Approver'] = '-'

    flight_data = pd.read_excel(flight_data_file_path)
    for index, row in clh.iterrows():
        flight_number_clh = row['Flight']
        matching_row = flight_data[flight_data['Flight Number'] == flight_number_clh]

        if not matching_row.empty:
            clh.at[index, 'ETD'] = matching_row['STD'].iloc[0]
            clh.at[index, 'ETA'] = matching_row['STA'].iloc[0]

    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Indigo')

    #Surya

    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    rows_deleted=[]
    mapping = {
        "COCHIN": "COK",
        "PUNE": "PNQ",
        "DELHI": "DEL",
        "BENGALURU": "BLR",
        "CHENNAI": "MAA",
        "MUMBAI": "BOM",
        "AHMEDABAD": "AMD",
        "KOLKATA": "CCU",
        "LUCKNOW": "LKO",
        "HYDERABAD": "HYD"
    }
    def map_airline(flight):
        airline_mapping = {
            "6E": "Indigo Air",
            "UK": "Vistara",
            "QP": "Akasa",
            "AI": "Air India",
            "IX": "Air India Express",
            "I5": "AirAsia",
            "9I": "Alliance Air",
            "SG": "Spicejet"
        }
        flight_str = str(flight)
        prefix = flight[:2] if len(flight_str) >= 2 else "" 
        return airline_mapping.get(prefix, "Unknown")


    file_path = r'C:\Project\CLH_Data\Surya.xlsx'
    surya = pd.read_excel(file_path)
    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh= pd.DataFrame(columns=columns)
    surya = pd.read_excel(file_path, skiprows=5, header=None)
    new_headers = [
        "AWB No", "AWB Date", "Department Name", "Consignor", "Consignee",
        "Origin", "Destination", "Flight #", "Pcs", "Act. Weight", "Charge Weight",
        "Status", "Delivered On", "Received By"
    ]
    surya.columns = new_headers
    surya = surya.iloc[:-1]
    surya = surya[surya['Department Name']!="AMAZON SELLER SERVICES PVT LTD-STUDIO"]
    surya.reset_index(inplace=True)

    clh['Date'] = pd.to_datetime(surya['AWB Date'], dayfirst=True).dt.strftime('%m/%d/%Y')
    clh['Origin'] = surya['Origin']
    clh['Dest'] = surya['Destination']
    clh['Origin'].replace(mapping, inplace=True)
    clh['Dest'].replace(mapping, inplace=True)
    clh['Lane'] = clh['Origin'] + '-' + clh['Dest']
    clh['Co-Loader'] = 'Surya Cargo Forwarders Pvt Ltd' 
    clh['Flight'] = surya['Flight #'].astype(str).str.replace("-", "").str.replace(" ", "")
    clh['AWBnum'] = surya['AWB No'].astype(str)  
    clh['CDnum'] = surya['AWB No']  
    clh['Pcs'] = surya['Pcs']  
    clh['Gross weight'] = surya['Act. Weight']  
    clh['Dim weight'] = surya['Charge Weight']  
    clh['Chg'] = surya['Charge Weight']  
    clh['Approver'] = '-'
    clh['Mode'] ='Dedicated'
    clh['Airline'] = clh['Flight'].astype(str).apply(map_airline)


    flight_data = pd.read_excel(flight_data_file_path)
    for index, row in clh.iterrows():
        flight_number_clh = row['Flight']
        matching_row = flight_data[flight_data['Flight Number'] == flight_number_clh]

        if not matching_row.empty:
            clh.at[index, 'ETD'] = matching_row['STD'].iloc[0]
            clh.at[index, 'ETA'] = matching_row['STA'].iloc[0]
    for index, row in clh.iterrows():
        if isinstance(row['AWBnum'], str) and len(row['AWBnum']) > 8:
            rows_deleted.append(index)
        if pd.isna(row['AWBnum']) or row['AWBnum'] == '':
            rows_deleted.append(index)

    wb = load_workbook(clh_file_path)
    if 'Surya' not in wb.sheetnames:
        wb.create_sheet('Surya')

    ws = wb['Surya']
    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Surya')



    #Spicejet
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook

    def map_airline(flight):
        if flight.startswith("SG7"):
            freighter_q400_flights = {
                "SG7083", "SG7084", "SG7401", "SG7402", "SG7404", "SG7405", 
                "SG7445", "SG7446", "SG7615", "SG7616", "SG7661", "SG7662", 
                "SG7667", "SG7668", "SG7669", "SG7441", "SG7442", "SG7617", 
                "SG7663", "SG7664", "SG7429", "SG7430"
            }
            freighter_cos_flights = {
                "SG7675", "SG7676", "SG7678", "SG7679", "SG7543", "SG7544"
            }
            amazon_air_flights = {
                "SG7612", "SG7613", "SG7613", "SG7614"
            }

            if flight in freighter_q400_flights:
                return "SpiceJet Freighter Q400"
            elif flight in freighter_cos_flights:
                return "SpiceJet Freighter COS"
            elif flight in amazon_air_flights:
                return "Amazon Air"
            else:
                return "SpiceJet Freighter"

        return "SpiceJet"



    spicejet = pd.read_excel(spicejet_file)
    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh = pd.DataFrame(columns=columns)

    clh['Date'] = pd.to_datetime(spicejet['FirstFlightDate'], dayfirst=True).dt.strftime('%m/%d/%Y')
    clh['Co-Loader']='Spicejet'
    clh['Origin'] = spicejet['OriginCode']
    clh['Dest'] = spicejet['DestinationCode']
    clh['Lane'] = clh['Origin'] + '-' + clh['Dest']
    clh['Flight'] = spicejet['FirstFlight'].apply(lambda x: x[:2] + '0' + x[2:] if len(x) == 5 else x)
    clh['Flight'] = clh['Flight'].astype(str)
    clh['Airline'] = clh['Flight'].apply(map_airline) 
    clh['AWBnum'] = spicejet['AWBNumber']  
    clh['CDnum'] = spicejet['AWBNumber']  
    clh['Pcs'] = spicejet['Pieces']  
    clh['Gross weight'] = spicejet['GrossWeight']  
    clh['Dim weight'] = spicejet['ChargedWeight']  
    clh['Chg'] = spicejet['ChargedWeight']  
    clh['Approver'] = '-'
    clh['Mode'] ='Dedicated'

    flight_data = pd.read_excel(flight_data_file_path)
    for index, row in clh.iterrows():
        flight_number_clh = row['Flight']
        matching_row = flight_data[flight_data['Flight Number'] == flight_number_clh]

        if not matching_row.empty:
            clh.at[index, 'ETD'] = matching_row['STD'].iloc[0]
            clh.at[index, 'ETA'] = matching_row['STA'].iloc[0]

    wb = load_workbook(clh_file_path)
    if 'Spicejet' not in wb.sheetnames:
        wb.create_sheet('Spicejet')

    ws = wb['Spicejet']
    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Spicejet')


    #POBC
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook

    def map_airline(flight):
        airline_mapping = {
            "6E": "Indigo Air",
            "UK": "Vistara",
            "QP": "Akasa",
            "AI": "Air India",
            "IX": "Air India Express",
            "I5": "AirAsia",
            "9I": "Alliance Air",
            "SG": "Spicejet"
        }
        prefix = str(flight)[:2] 
        return airline_mapping.get(prefix, "Surface")

    pobc = pd.read_excel(pobc_file,header=2)
    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh = pd.DataFrame(columns=columns)

    clh['Date'] = pd.to_datetime(pobc['Date'], dayfirst=True).dt.strftime('%m/%d/%Y')
    clh['Origin'] = pobc['Origin']
    clh['Dest'] = pobc['Dest']
    clh['Lane'] = clh['Origin'] + '-' + clh['Dest']
    clh['Co-Loader'] = 'Patel On-Board Couriers Ltd' 
    clh['Flight'] = pobc['Flt No'].str.replace("-", "").str.replace(" ", "").str.replace(".","")
    clh['Flight'] = clh['Flight'].apply(lambda x: x[:2] + '0' + x[2:] if isinstance(x, str) and len(x) == 5 else x)
    clh['AWBnum'] = pobc['Mawb No.']  
    clh['CDnum'] = pobc['CD NO']  
    clh['Pcs'] = pobc['Bags']  
    clh['Gross weight'] = pobc['Gross Wght']  
    clh['Dim weight'] = pobc['Dim Wght']  
    clh['Chg'] = pobc['Chg Wght']  
    clh['Approver'] =pobc['Approver Name']
    clh['Mode'] =pobc['Mode']
    clh['Airline'] = clh['Flight'].apply(map_airline)
    mask = clh['Airline'].str.contains('Surface', case=False)
    clh.loc[mask, 'Chg'] = 0
    clh.loc[mask, 'Dim weight'] = 0
    clh.loc[mask,'ETD']=0
    clh.loc[mask,'ETA']=0


    flight_data = pd.read_excel(flight_data_file_path)
    for index, row in clh.iterrows():
        flight_number_clh = row['Flight']
        matching_row = flight_data[flight_data['Flight Number'] == flight_number_clh]

        if not matching_row.empty:
            clh.at[index, 'ETD'] = matching_row['STD'].iloc[0]
            clh.at[index, 'ETA'] = matching_row['STA'].iloc[0]
            #clh.at[index, 'Airline'] = matching_row['Airline'].iloc[0]
    wb = load_workbook(clh_file_path)
    if 'Pobc' not in wb.sheetnames:
        wb.create_sheet('Pobc')

    ws = wb['Pobc']
    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Pobc')

    #KRBL
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook


    def map_airline(flight):
        airline_mapping = {
            "6E": "Indigo Air",
            "UK": "Vistara",
            "QP": "Akasa",
            "AI": "Air India",
            "IX": "Air India Express",
            "I5": "AirAsia",
            "9I": "Alliance Air",
            "SG": "Spicejet"
        }
        prefix = flight[:2]  
        return airline_mapping.get(prefix, "Unknown")


    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh = pd.DataFrame(columns=columns)



    krbl = pd.read_excel(krbl_file,header=1)
    krbl['Flt. Date']= krbl['Flt. Date'].astype(str).apply(lambda x: x[:-4] + '2024' if x[-4:] == '0224' else x)
    clh['Date'] = pd.to_datetime(krbl['Flt. Date'], dayfirst=True, errors='coerce')
    clh['Date'] = clh['Date'].dt.strftime('%m/%d/%Y')
    clh['Origin'] =krbl['Origin']
    clh['Dest'] =krbl['Destination']
    clh['Lane'] =clh['Origin'] + '-' + clh['Dest']
    clh['Flight'] =krbl['Planned Flt'].str.replace("-", "").str.replace(" ", "")
    clh['Flight'] =clh['Flight'].apply(lambda x: x[:2] + '0' + x[2:] if isinstance(x, str) and len(x) == 5 else x)
    clh['AWBnum'] =krbl['AWB No.'].apply(lambda x: x.split('-')[1] if isinstance(x, str) and '-' in x else x) 
    clh['CDnum'] =clh['AWBnum']  
    clh['Pcs'] = krbl['Pcs']  
    clh['Airline'] = clh['Flight'].apply(map_airline)
    clh['Gross weight'] = krbl['G/Wt']  
    clh['Dim weight'] = krbl['C/Wt']  
    clh['Chg'] = krbl['C/Wt']  
    clh['Co-Loader']='Kreative Rainbowbliss Logistics'
    clh['Approver'] ='-'
    clh['Mode'] ='Dedicated'

    flight_data = pd.read_excel(flight_data_file_path)
    for index, row in clh.iterrows():
        flight_number_clh = row['Flight']
        matching_row = flight_data[flight_data['Flight Number'] == flight_number_clh]

        if not matching_row.empty:
            clh.at[index, 'ETD'] = matching_row['STD'].iloc[0]
            clh.at[index, 'ETA'] = matching_row['STA'].iloc[0]


    wb = load_workbook(clh_file_path)
    if 'Krbl' not in wb.sheetnames:
        wb.create_sheet('Krbl')

    ws = wb['Krbl']
    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Krbl')

    #KRBL-airasia
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook


    def map_airline(flight):
        airline_mapping = {
            "6E": "Indigo Air",
            "UK": "Vistara",
            "QP": "Akasa",
            "AI": "Air India",
            "IX": "Air India Express",
            "I5": "AirAsia",
            "9I": "Alliance Air",
            "SG": "Spicejet"
        }
        prefix = flight[:2]  
        return airline_mapping.get(prefix, "Unknown")


    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh = pd.DataFrame(columns=columns)



    krbla = pd.read_excel(krbla_file,header=1)
    clh['Date'] = krbla['Flt. Date']
    clh['Date']=pd.to_datetime(clh['Date'], dayfirst=True, errors='coerce')
    clh['Date'] = clh['Date'].dt.strftime('%m/%d/%Y')
    clh['Origin'] =krbla['Origin']
    clh['Dest'] =krbla['Destination']
    clh['Lane'] =clh['Origin'] + '-' + clh['Dest'] 
    clh['Flight'] =krbla['Planned Flt'].str.replace("-", "").str.replace(" ", "")
    clh['Flight'] =clh['Flight'].apply(lambda x: x[:2] + '0' + x[2:] if isinstance(x, str) and len(x) == 5 else x)
    clh['AWBnum'] =krbla['AWB NO.'].apply(lambda x: x.split('-')[1] if isinstance(x, str) and '-' in x else x) 
    clh['CDnum'] =clh['AWBnum']  
    clh['Pcs'] = krbla['Pcs']  
    clh['Airline'] = 'AirAsia India'
    clh['Gross weight'] = krbla['G/Wt']  
    clh['Dim weight'] = krbla['C/Wt']  
    clh['Chg'] = krbla['C/Wt']  
    clh['Co-Loader']='Kreative Rainbowbliss Logistics'
    clh['Approver'] ='-'
    clh['Mode'] = 'Dedicated'
    flight_data = pd.read_excel(flight_data_file_path)
    for index, row in clh.iterrows():
        flight_number_clh = row['Flight']
        matching_row = flight_data[flight_data['Flight Number'] == flight_number_clh]

        if not matching_row.empty:
            clh.at[index, 'ETD'] = matching_row['STD'].iloc[0]
            clh.at[index, 'ETA'] = matching_row['STA'].iloc[0]
    wb = load_workbook(clh_file_path)
    if 'Krbla' not in wb.sheetnames:
        wb.create_sheet('Krbla')

    ws = wb['Krbla']
    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Krbla')

    #Index
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook

    mapping = {
        "COCHIN": "COK",
        "PUNE": "PNQ",
        "DELHI": "DEL",
        "BENGALURU": "BLR",
        "CHENNAI": "MAA",
        "MUMBAI": "BOM",
        "AHMEDABAD": "AMD",
        "KOLKATA": "CCU",
        "LUCKNOW": "LKO",
        "HYDERABAD": "HYD"
    }
    def map_airline(flight):
        airline_mapping = {
            "6E": "Indigo Air",
            "UK": "Vistara",
            "QP": "Akasa",
            "AI": "Air India",
            "IX": "Air India Express",
            "I5": "AirAsia",
            "9I": "Alliance Air",
            "SG": "Spicejet"
        }
        flight_str = str(flight)
        prefix = flight[:2] if len(flight_str) >= 2 else "" 
        return airline_mapping.get(prefix, "Unknown")


    ind = pd.read_excel(index_file,sheet_name="Sheet1")
    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh = pd.DataFrame(columns=columns)

    clh['Date'] = ind['BOOKING DATE']
    clh['Date'] = pd.to_datetime(clh['Date'], format='%d.%m.%Y')
    clh['Date'] = clh['Date'].dt.strftime('%m/%d/%Y')
    clh['Origin'] = ind['ORIGIN']
    clh['Dest'] = ind['DEST.']
    clh['Origin'].replace(mapping, inplace=True)
    clh['Dest'].replace(mapping, inplace=True)
    clh['Lane'] = clh['Origin'] + '-' + clh['Dest']
    clh['Co-Loader'] = 'Index' 
    clh['Flight'] = ind['FLIGHT NO.'].str.replace("-", "").str.replace(" ", "")
    clh['Flight'] =clh['Flight'].apply(lambda x: x[:2] + '0' + x[2:] if isinstance(x, str) and len(x) == 5 else x)
    clh['AWBnum'] = ind['CD/AWB']  
    clh['CDnum'] = ind['CD/AWB']  
    clh['Pcs'] = ind['PCS']  
    clh['Gross weight'] = ind['ACT. WEIGHT']  
    clh['Dim weight'] = ind['CH. WEIGHT']  
    clh['Chg'] = ind['CH. WEIGHT']  
    clh['Approver'] = '-'
    clh['Mode'] ='Dedicated'
    clh['Airline'] = clh['Flight'].astype(str).apply(map_airline)

    flight_data = pd.read_excel(flight_data_file_path)
    for index, row in clh.iterrows():
        flight_number_clh = row['Flight']
        matching_row = flight_data[flight_data['Flight Number'] == flight_number_clh]

        if not matching_row.empty:
            clh.at[index, 'ETD'] = matching_row['STD'].iloc[0]
            clh.at[index, 'ETA'] = matching_row['STA'].iloc[0]
    wb = load_workbook(clh_file_path)
    if 'Index' not in wb.sheetnames:
        wb.create_sheet('Index')

    ws = wb['Index']
    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Index')

    #Labriynth-Sea
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    import datetime
    def calculate_gross_weight(origin, pcs):
        if origin in ["MAA", "IXZ"]:
            return pcs * 1000000 / 6000
        else:
            return None

    current_date = datetime.date.today()

    week_number = str(current_date.isocalendar()[1])
    current_year = str(current_date.year)
    sheet_name = "Week " + week_number + " " + current_year
    lab = pd.read_excel(labriynth_1_file,sheet_name="March")
    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh = pd.DataFrame(columns=columns)

    clh['Date'] = lab['Depart Date']
    clh['Date'] = pd.to_datetime(clh['Date'], format='%d.%m.%Y')
    clh['Date'] = clh['Date'].dt.strftime('%m/%d/%Y')
    clh['Origin'] = lab['ORIGIN']
    clh['Dest'] = lab['DEST'].str[:3]
    clh['Lane'] = clh['Origin'] + '-' + clh['Dest']
    clh['Co-Loader'] = 'Labyrinth Logistics Pvt Ltd' 
    clh['Flight'] = 'Sea'
    clh['AWBnum'] = lab['Cointainer No']  
    clh['CDnum'] = lab['Cointainer No']  
    clh['Pcs'] = lab['CBM']  
    clh['Gross weight'] = lab.apply(lambda row: calculate_gross_weight(row['ORIGIN'], row['CBM']), axis=1)
    clh['Dim weight'] = clh['Gross weight']
    clh['Chg'] = clh['Gross weight']  
    clh['Approver'] = '-'
    clh['Mode'] ='Dedicated'
    clh['ETA'] = pd.to_datetime('0:00', format='%H:%M').time()
    clh['ETD'] = pd.to_datetime('0:00', format='%H:%M').time()

    wb = load_workbook(clh_file_path)
    if 'Labriynth' not in wb.sheetnames:
        wb.create_sheet('Labriynth')

    ws = wb['Labriynth']
    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Labriynth')

    #Labriynth-RL sea
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from datetime import datetime, timedelta
    import datetime
    def calculate_gross_weight(origin, pcs):
        if origin in ["MAA", "IXZ"]:
            return pcs * 1000000 / 6000
        else:
            return None

    first_day_of_month = current_date.replace(day=1)
    a1 = str(first_day_of_month.day).zfill(2)  
    a2 = str(first_day_of_month.month).zfill(2)  
    a3 = str(first_day_of_month.year)
    previous_saturday = current_date - timedelta(days=(current_date.weekday() + 2) % 7)
    b1 = str(previous_saturday.day).zfill(2) 
    b2 = str(previous_saturday.month).zfill(2)  
    b3 = str(previous_saturday.year)
    a = a2 + "/" + a1 + "/" + a3
    b = b2 + "/" + b1 + "/" + b3


    current_date = datetime.date.today()
    week_number = str(current_date.isocalendar()[1])
    current_year = str(current_date.year)
    sheet_name = "Week " + week_number + " " + current_year
    lab = pd.read_excel(labriynth_file,sheet_name="Sea RL")
    columns = ['Date', 'Origin', 'Dest', 'Lane', 'Co-Loader', 'Airline', 'Flight', 'ETD', 'ETA', 'Mode', 'Approver', 'AWBnum', 'CDnum', 'Pcs', 'Gross weight', 'Dim weight', 'Chg']
    clh = pd.DataFrame(columns=columns)

    clh['Date'] = lab['ETD']
    clh['Date'] = pd.to_datetime(clh['Date'], format='%d.%m.%Y')
    clh['Date'] = clh['Date'].dt.strftime('%m/%d/%Y')
    clh['Origin'] = lab['ORIGIN'].str[:3]
    clh['Dest'] = lab['DEST']
    clh['Lane'] = clh['Origin'] + '-' + clh['Dest']
    clh['Co-Loader'] = 'Labyrinth Logistics Pvt Ltd' 
    clh['Flight'] = 'Sea'
    clh['AWBnum'] = lab['Cointainer No']  
    clh['CDnum'] = lab['Cointainer No']  
    clh['Pcs'] = lab['CBM']  
    clh['Gross weight'] = np.where((clh['Origin'] == 'MAA') | (clh['Origin'] == 'IXZ'), clh['Pcs'] * 1000000 / 6000, clh['Pcs'])
    clh['Dim weight'] = clh['Gross weight']
    clh['Chg'] = clh['Gross weight']  
    clh['Approver'] = '-'
    clh['Mode'] ='Dedicated'
    clh['ETD'] = pd.to_datetime('0:00', format='%H:%M').time()
    clh['ETA'] = pd.to_datetime('0:00', format='%H:%M').time()
    clh = clh[(clh['Date'] >= a) & (clh['Date'] <= b)]

    wb = load_workbook(clh_file_path)
    if 'Labriynth-R' not in wb.sheetnames:
        wb.create_sheet('Labriynth-R')

    ws = wb['Labriynth-R']
    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        clh.to_excel(writer, index=False, sheet_name='Labriynth-R')

    #This logic checks the number of rows
    import pandas as pd


    xls = pd.ExcelFile(clh_file_path)
    row_counts = {}
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(clh_file_path, sheet_name=sheet_name)
        row_counts[sheet_name] = len(df)
    for sheet_name, count in row_counts.items():
        print(f"Sheet '{sheet_name}' has {count} rows.")

    #Appends all data to one sheet named CLH and changes the column names
    import pandas as pd
    from openpyxl.styles import NamedStyle


    xls = pd.ExcelFile(clh_file_path)
    dfs = {sheet_name: xls.parse(sheet_name) for sheet_name in xls.sheet_names}
    combined_data = pd.concat(dfs.values(), ignore_index=True)
    column_replacements = {
        'Dest': 'Destination',
        'Flight': 'FlightNo.',
        'ETD': 'ATD',
        'ETA': 'ATA',
        'AWBnum': 'MawbNo.',
        'CDnum': 'CD No.',
        'Pcs': 'Pcs.',
        'Gross weight': 'Gross Wght',
        'Dim weight': 'Dim',
        'Chargeable weight': 'Chg'
    }


    combined_data.rename(columns=column_replacements, inplace=True)
    combined_data.dropna(subset=['Date'], inplace=True)
    combined_data.reset_index(drop=True, inplace=True)
    combined_data['MawbNo.'] = combined_data['MawbNo.'].astype(str)    
    #combined_data['ATD'] = pd.to_datetime(combined_data['ATD'], errors='coerce')
    #combined_data['ATA'] = pd.to_datetime(combined_data['ATA'], errors='coerce')
    ws = wb.active
    num_rows = df.shape[0]
    integer_style = NamedStyle(name='integer')
    integer_style.number_format = '0'
    '''
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=num_rows):
        for cell in row:
            cell.style = integer_style
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=num_rows):
        for cell in row:
            cell.style = integer_style
    wb.save(clh_file_path)
    '''
    with pd.ExcelWriter(clh_file_path, mode='a', engine='openpyxl') as writer:
        combined_data.to_excel(writer, index=False, sheet_name='CLH', engine='xlsxwriter')

    # Missing Flight data

    import pandas as pd

    def check_flight_numbers(clh_file_path, flight_data_file_path):
        clh_df = pd.read_excel(clh_file_path, sheet_name='CLH')

        # Read Flight data file
        flight_data_df = pd.read_excel(flight_data_file_path)

        # Perform check and create status column
        clh_df['Status'] = clh_df['FlightNo.'].isin(flight_data_df['Flight Number'])

        # Find missing flights excluding NaN and 'Sea' values
        missing_flights = clh_df.loc[~clh_df['Status'] & ~clh_df['FlightNo.'].isna() & (clh_df['FlightNo.'] != 'Sea'), 'FlightNo.']

        # Count missing flights
        missing_flights_count = len(missing_flights)

        # Open CLH inputs file for writing
        with pd.ExcelWriter(clh_file_path, engine='openpyxl', mode='a') as writer:
            # Remove existing 'CLH' sheet if it exists
            if 'CLH' in writer.sheets:
                writer.book.remove(writer.sheets['CLH'])
            # Write the CLH dataframe to a new 'CLH' sheet
            clh_df.to_excel(writer, sheet_name='CLH', index=False)

        # Display check completion message
        print("Check completed.")
        print(f"Number of missing flight information: {missing_flights_count}")
        print("Missing flights:")
        print(missing_flights.to_string(index=False))

    # Replace clh_file_path and flight_data_file_path with your file paths
    check_flight_numbers(clh_file_path, flight_data_file_path)

    # Cost Calculation

    import numpy as np

    import pandas as pd

    from datetime import datetime as dt

    import pandas as pd
    clh=pd.read_excel(clh_file_path,sheet_name='CLH')
    failed_df = pd.DataFrame()

    # INDIGO

    indigo=clh[clh['Co-Loader']=='IndiGo Air']

    indigo.head()

    indigo['flt_type']=''

    #prime/gcr classification
    indigo.reset_index(drop=True, inplace=True)
    from datetime import time
    #indigo['ATD'] = pd.to_datetime(indigo['ATD'])
    #indigo['ATD'] = indigo['ATD'].apply(lambda x: x.time())
    indigo['ATD'] = pd.to_datetime(indigo['ATD'], format='%H:%M:%S')

    for i in range(len(indigo)):
        if indigo['ATD'][i].time()>=time(2,0,0) and indigo['ATD'][i].time()<=time(9,0,0):
            indigo['flt_type'][i]='prime'
        else:
            indigo['flt_type'][i]='non-prime'
    indigo.head()

    indigo['flt_rate']=''

    #flight cost
    for i in range(len(indigo)):
        if indigo['Mode'][i]=='Express':
            indigo['flt_rate'][i]=105
        else:
            if indigo['flt_type'][i]=='prime':
                indigo['flt_rate'][i]=39
            else:
                indigo['flt_rate'][i]=25.5

    #GAU rates
    gau_rates=pd.read_excel(rates_file,sheet_name='indigo-gau')
    for i in range(len(indigo)):
        for j in range(len(gau_rates)):
            if indigo['Lane'][i]==gau_rates['lane'][j]:
                if indigo['Chg'][i]<45:
                    indigo['flt_rate'][i]=gau_rates['n_rate'][j]
                elif indigo['Chg'][i]<100:
                    indigo['flt_rate'][i]=gau_rates['+45'][j]
                elif indigo['Chg'][i]<300:
                    indigo['flt_rate'][i]=gau_rates['+100'][j]
                elif indigo['Chg'][i]<500:
                    indigo['flt_rate'][i]=gau_rates['+300'][j]
                elif indigo['Chg'][i]<1000:
                    indigo['flt_rate'][i]=gau_rates['+500'][j]
                else:
                    indigo['flt_rate'][i]=gau_rates['+1000'][j]


    indigo['flt_cost']=''
    for i in range(len(indigo)):
        indigo['flt_cost'][i]=max(indigo['Chg'][i]*indigo['flt_rate'][i],1500)

    #security cost
    indigo['security']=1*indigo['Chg']

    indigo['admin_cost']=''

    #admin cost
    metro=['DEL','BOM','BLR','MAA','CCU','HYD']
    for i in range(len(indigo)):
        if indigo['Origin'][i] in metro:
            indigo['admin_cost'][i]=3*indigo['Chg'][i]
        else:
            indigo['admin_cost'][i]=2*indigo['Chg'][i]

    cut=pd.read_excel(rates_file,sheet_name='indigo-misc')

    #CUT_OUTBOUND_COST
    indigo['cut_ob']=''
    for i in range(len(indigo)):
        outbound=0
        for j in range(len(cut)):
            if cut['station'][j]==indigo['Origin'][i]:
                outbound=max(indigo['Chg'][i]*cut['OB_rate'][j],cut['OB_min'][j])+outbound
        indigo['cut_ob'][i]=outbound

    #CUT_INBOUND_COST
    indigo['cut_ib']=''
    for i in range(len(indigo)):
        inbound=0
        for j in range(len(cut)):
            if cut['station'][j]==indigo['Destination'][i]:
                inbound=max(indigo['Chg'][i]*cut['IB_rate'][j],cut['IB_min'][j])+inbound
        indigo['cut_ib'][i]=inbound

    indigo['cut_total']=indigo['cut_ob']+indigo['cut_ib']

    indigo['net_cost']=indigo['flt_cost']+indigo['security']+indigo['admin_cost']+indigo['cut_total']

    indigo.head(10)

    # Spicejet

    spicejet=clh[clh['Co-Loader']=='Spicejet']
    spicejet=spicejet.reset_index(drop=True)

    metro=['DEL','MAA','BOM','BLR']
    spicejet['category']=''
    Rates = pd.read_excel(rates_file,sheet_name='sg')
    prime_flts = Rates['key'].tolist()
    prime_flts = [lane.replace(" ", "") for lane in prime_flts]
    spicejet['flt_rate']=''
    spicejet['category']=''
    category_rate_mapping = {
        'metro to metrogcr': 26,
        'non metro to metrogcr': 22,
        'metro to non metrogcr': 22,
        'non metro to non metrogcr': 22
    }
    for i in range(len(spicejet)):
        if spicejet['Origin'][i] in metro:
            if spicejet['Destination'][i] in metro:
                spicejet['category'][i]= 'metro to metro'
            else:
                spicejet['category'][i]= 'metro to non metro'
        else:
            if spicejet['Destination'][i] in metro:
                spicejet['category'][i]= 'non metro to metro'
            else:
                spicejet['category'][i]= 'non metro to non metro'
    for i in range(len(spicejet)):
        origin_dest_match = False
        flight_number_match = False

        for lane in prime_flts:
            if (spicejet['Origin'][i] + spicejet['Destination'][i]) == lane[:6]:
                origin_dest_match = True
                flight_number = lane[6:]

                if flight_number == spicejet['FlightNo.'][i]:
                    flight_number_match = True

                    rate_row = Rates[Rates['flights'] == flight_number]

                    if not rate_row.empty:
                        if spicejet['Chg'][i] < 45:
                            spicejet['flt_rate'][i] = rate_row['N Rate'].values[0]
                        elif spicejet['Chg'][i] < 100:
                            spicejet['flt_rate'][i] = rate_row['+45'].values[0]
                        else:
                            spicejet['flt_rate'][i] = rate_row['+100'].values[0]

                    spicejet['category'][i] = spicejet['category'][i] + 'prime'
                    break

        if origin_dest_match and not flight_number_match:
            for lane in prime_flts:
                if (spicejet['Origin'][i] + spicejet['Destination'][i]) == lane[:6]:
                    flight_number = lane[6:]

                    if flight_number == 'all_flights':
                        key_to_match = spicejet['Origin'][i] + spicejet['Destination'][i] + 'all_flights'

                        rate_row_all_flights = Rates[Rates['key'] == key_to_match]

                        if not rate_row_all_flights.empty:
                            if spicejet['Chg'][i] < 45:
                                spicejet['flt_rate'][i] = rate_row_all_flights['N Rate'].values[0]
                            elif spicejet['Chg'][i] < 100:
                                spicejet['flt_rate'][i] = rate_row_all_flights['+45'].values[0]
                            else:
                                spicejet['flt_rate'][i] = rate_row_all_flights['+100'].values[0]

                        spicejet['category'][i] = spicejet['category'][i] + 'prime'
                        break

        if not origin_dest_match or (origin_dest_match and not flight_number_match):
            spicejet['category'][i] = spicejet['category'][i] + 'gcr'

    for i, row in spicejet.iterrows():
        category = row['category']
        if category in category_rate_mapping:
            spicejet.at[i, 'flt_rate'] = category_rate_mapping[category]


    spicejet['flt_cost']=''

    for i in range(len(spicejet)):
        spicejet['flt_cost'][i] = max(spicejet['flt_rate'][i] * spicejet['Chg'][i], 1600)
    spicejet.head(10)     

    #fuel surcharge
    ob_cities=['BOM','AMD','CCU','DEL','HYD','MAA','PNQ','BLR']
    for i in range(len(spicejet)):
        if spicejet['Airline'][i]=='SpiceJet' and spicejet['Origin'][i] in ob_cities:
            spicejet["fuel_surcharge"]=3*spicejet['Chg']
        else:
            spicejet["fuel_surcharge"]=2*spicejet['Chg']

    #security
    spicejet['security']=1.5*spicejet['Chg']

    #CUT_OUTBOUND_COST
    cut=pd.read_excel(rates_file,sheet_name='sg-misc')
    spicejet['cut_ob']=''
    for i in range(len(spicejet)):
        outbound=0
        for j in range(len(cut)):
            if cut['station'][j]==spicejet['Origin'][i]:
                outbound=max(spicejet['Chg'][i]*cut['OB_rate'][j],cut['OB_min'][j])+outbound
            spicejet['cut_ob'][i]=outbound

    #CUT_INBOUND_COST
    spicejet['cut_ib']=''
    for i in range(len(spicejet)):
        inbound=0
        for j in range(len(cut)):
            if cut['station'][j]==spicejet['Destination'][i]:
                inbound=max(spicejet['Chg'][i]*cut['IB_rate'][j],cut['IB_min'][j])+inbound
        spicejet['cut_ib'][i]=inbound

    #Sector/Flight Surcharge
    surcharge=pd.read_excel(rates_file,sheet_name='sg-surcharge')
    spicejet['surcharge']=0
    for i in range(len(spicejet)):
        for j in range(len(surcharge)):
            if surcharge['flights'][j]=='all_flights':
                if spicejet['Lane'][i]==surcharge['lane'][j]:
                    spicejet['surcharge'][i]=surcharge['surcharge'][j]
        for j in range(len(surcharge)):
                if spicejet['Lane'][i]==surcharge['lane'][j] and spicejet['FlightNo.'][i]==surcharge['flights'][j]:
                    spicejet['surcharge'][i]=surcharge['surcharge'][j]

    spicejet['awb_charges']=150
    spicejet['do_charges']=300

    spicejet['net_cost']=spicejet['flt_cost']+spicejet['fuel_surcharge']+spicejet['security']+(spicejet['surcharge']*spicejet['Chg'])+spicejet['cut_ob']+spicejet['cut_ib']+spicejet['awb_charges']+spicejet['do_charges']

    spicejet.head(5)

    # KRBL

    #AIRASIA

    airasia=clh[(clh['Co-Loader']=='Kreative Rainbowbliss Logistics') & (clh['Airline']=='AirAsia India')]
    airasia=airasia.reset_index(drop=True)

    airasia.head(5)

    airasia_rates=pd.read_excel(rates_file,sheet_name='airasia_rates')

    airasia['flt_rates']=''
    for i in range(len(airasia)):
        for j in range(len(airasia_rates)):
            if (airasia['Origin'][i]+airasia['Destination'][i]+airasia['FlightNo.'][i])==airasia_rates['key'][j]:
                airasia['flt_rates'][i]=airasia_rates['rate'][j]

    airasia.head(5)

    airasia['flt_rates']=airasia['flt_rates'].apply(pd.to_numeric)

    airasia['flt_cost']=''
    flag=0
    for i in range(len(airasia)):
        try:
            flag=1
            print(flag)
            airasia['flt_cost'][i] = max(airasia['flt_rates'][i] * airasia['Chg'][i], 1400)
        except Exception as e:
            flag=2
            print(flag)


    #CUT_OUTBOUND_COST
    cut=pd.read_excel(rates_file,sheet_name='airasia-misc')
    airasia['cut_ob']=''
    for i in range(len(airasia)):
        outbound=0
        for j in range(len(cut)):
            if cut['station'][j]==airasia['Origin'][i]:
                outbound=max(airasia['Chg'][i]*cut['OB_rate'][j],cut['OB_min'][j])+outbound
            airasia['cut_ob'][i]=outbound

    #CUT_INBOUND_COST
    airasia['cut_ib']=''
    for i in range(len(airasia)):
        inbound=0
        for j in range(len(cut)):
            if cut['station'][j]==airasia['Destination'][i]:
                inbound=max(airasia['Chg'][i]*cut['IB_rate'][j],cut['IB_min'][j])+inbound
        airasia['cut_ib'][i]=inbound

    airasia['AWB+DO']=600

    airasia['KRBL_handling']=2*airasia['Chg']

    airasia['misc']=3*airasia['Chg']

    airasia['net_cost']=''
    flag=0
    for i in range(len(airasia)):
        try:
            flag=1
            print(flag)
            airasia['net_cost'][i]=airasia['flt_cost'][i]+airasia['cut_ob'][i]+airasia['cut_ib'][i]+airasia['AWB+DO'][i]+airasia['KRBL_handling'][i]+airasia['misc'][i]
        except Exception as e:
            flag=2
            #print(flag)
            continue

    #Vistara

    vistara=clh[(clh['Co-Loader']=='Kreative Rainbowbliss Logistics') & (clh['Airline']=='Vistara')]
    vistara=vistara.reset_index(drop=True)

    vistara['flt_rates']=''
    vistara_rates=pd.read_excel(rates_file,sheet_name='vistara')
    for i in range(len(vistara)):
        for j in range(len(vistara_rates)):            
            if vistara_rates['flights'][j]=='all_flights':
                if vistara['Lane'][i]==vistara_rates['lane'][j]:
                    if vistara['Chg'][i]<45:
                        vistara['flt_rates'][i]=vistara_rates['n_rate'][j]
                    elif vistara['Chg'][i]<100:
                        vistara['flt_rates'][i]=vistara_rates['+45kgs'][j]
                    elif vistara['Chg'][i]>=100:
                        vistara['flt_rates'][i]=vistara_rates['+100kgs'][j]
        for j in range(len(vistara_rates)):
            if vistara['Lane'][i]==vistara_rates['lane'][j] and vistara['FlightNo.'][i]==vistara_rates['flights'][j]:
                    if vistara['Chg'][i]<45:
                        vistara['flt_rates'][i]=vistara_rates['n_rate'][j]
                    elif vistara['Chg'][i]<100:
                        vistara['flt_rates'][i]=vistara_rates['+45kgs'][j]
                    elif vistara['Chg'][i]>=100:
                        vistara['flt_rates'][i]=vistara_rates['+100kgs'][j]
    #in the above logic the model first adds rates for lane wise where specific flight rates are not given in rate card and then
    #adds the rates for specifc flights

    vistara['flt_cost']=''
    for i in range(len(vistara)):
        try:
            #failed_rows = vistara[(vistara['flt_rates'] == 0) | vistara['flt_rates'].isnull()][['FlightNo.', 'Co-Loader', 'Lane']]
            #failed_df = failed_df.append(failed_rows, ignore_index=True)
            vistara['flt_cost'][i] = max(vistara['flt_rates'][i] * vistara['Chg'][i], 1400)
        except Exception as e:
            continue

    #admin cost
    vistara['admin_cost']=''
    metro=['DEL','BOM','BLR','MAA','CCU','HYD','AMD']
    for i in range(len(vistara)):
        if vistara['Origin'][i] in metro:
            vistara['admin_cost'][i]=2*vistara['Chg'][i]
        else:
            vistara['admin_cost'][i]=1*vistara['Chg'][i]

    #sector/flight surcharge
    s={
        "lane":['DEL-MAA','DEL-IXL','IXC-BLR','DEL-PNQ','DEL-HYD','DEL-HYD'],
        "flight":['all_flights','all_flights','all_flights','UK0971','UK0829','UK0859'],
        "surcharge":[4,30,5,3,3,3]
    }

    vistara_surcharge=pd.DataFrame(s)
    vistara_surcharge

    #sector/flight surcharge
    vistara['surcharge']=0
    for i in range(len(vistara)):
        for j in range(len(vistara_surcharge)):
            if vistara_surcharge['flight'][j]=='all_flights':
                if vistara['Lane'][i]==vistara_surcharge['lane'][j]:
                    vistara['surcharge'][i]=vistara_surcharge['surcharge'][j]
        for j in range(len(vistara_surcharge)):
                if vistara['Lane'][i]==vistara_surcharge['lane'][j] and vistara['FlightNo.'][i]==vistara_surcharge['flight'][j]:
                    vistara['surcharge'][i]=vistara_surcharge['surcharge'][j]

    #CUT_OUTBOUND_COST
    cut=pd.read_excel(rates_file,sheet_name='vistara-misc')
    vistara['cut_ob']=''
    for i in range(len(vistara)):
        outbound=0
        for j in range(len(cut)):
            if cut['station'][j]==vistara['Origin'][i]:
                outbound=max(vistara['Chg'][i]*cut['OB_rate'][j],cut['OB_min'][j])+outbound
            vistara['cut_ob'][i]=outbound

    #CUT_INBOUND_COST
    vistara['cut_ib']=''
    for i in range(len(vistara)):
        inbound=0
        for j in range(len(cut)):
            if cut['station'][j]==vistara['Destination'][i]:
                inbound=max(vistara['Chg'][i]*cut['IB_rate'][j],cut['IB_min'][j])+inbound
        vistara['cut_ib'][i]=inbound

    vistara['handling_cost']=1.75*vistara['Chg']
    vistara['AWB+DO']=300+200

    vistara['net_cost']=''
    for i in range(len(vistara)):
        try:
            vistara['net_cost'][i]=vistara['flt_cost'][i]+vistara['admin_cost'][i]+vistara['surcharge'][i]*vistara['Chg'][i]+vistara['cut_ob'][i]+vistara['cut_ib'][i]+vistara['handling_cost'][i]+vistara['AWB+DO'][i]
        except Exception as e:
            continue


    #vistara['net_cost']=vistara['flt_cost']+vistara['admin_cost']+vistara['surcharge']*vistara['Chg']+vistara['cut_ob']+vistara['cut_ib']+vistara['handling_cost']+vistara['AWB+DO']

    vistara.head(5)

    #KRBL -> INDIGO and AKASA

    krbl=clh[(clh['Co-Loader']=='Kreative Rainbowbliss Logistics') & (clh['Airline']!='AirAsia India') &(clh['Airline']!='Vistara')]
    krbl=krbl.reset_index(drop=True)

    krbl_rates=pd.read_excel(rates_file,sheet_name='krbl')
    krbl['flt_rate']=''
    #krbl['FlightNo.']=krbl['FlightNo.'].astype(str)
    for i in range(len(krbl)):
        for j in range(len(krbl_rates)):
            if krbl['Origin'][i]+krbl['Destination'][i]+krbl['FlightNo.'][i]==krbl_rates['key'][j]:
                if krbl['Chg'][i]<300:
                    krbl['flt_rate'][i]=krbl_rates['+100'][j]
                elif krbl['Chg'][i]<500:
                    krbl['flt_rate'][i]=krbl_rates['+300'][j]
                else:
                    krbl['flt_rate'][i]=krbl_rates['+500'][j]
                break
            elif krbl['Origin'][i]+krbl['Destination'][i]+krbl['FlightNo.'][i][:2]+"9999"==krbl_rates['key'][j]:
                if pd.isnull(krbl_rates['start_time'][j]):
                    if krbl['Chg'][i]<300:
                        krbl['flt_rate'][i]=krbl_rates['+100'][j]
                    elif krbl['Chg'][i]<500:
                        krbl['flt_rate'][i]=krbl_rates['+300'][j]
                    else:
                        krbl['flt_rate'][i]=krbl_rates['+500'][j]
                break

    #Rates for flights with time constraints
    for j in range(len(krbl_rates)):
        for i in range(len(krbl)):
            if pd.isnull(krbl_rates['start_time'][j]) is False and krbl_rates['key'][j]==krbl['Origin'][i]+krbl['Destination'][i]+krbl['FlightNo.'][i][:2]+"9999":
                if krbl['ATD'][i]>=krbl_rates['start_time'][j] and krbl['ATD'][i]<=krbl_rates['end_time'][j]:
                    if krbl['Chg'][i]<300:
                        krbl['flt_rate'][i]=krbl_rates['+100'][j]
                    elif krbl['Chg'][i]<500:
                        krbl['flt_rate'][i]=krbl_rates['+300'][j]
                    else:
                        krbl['flt_rate'][i]=krbl_rates['+500'][j]

    krbl['flt_rate']=krbl['flt_rate'].apply(pd.to_numeric)

    krbl['net_cost'] = ''


    for i in range(len(krbl)):
        try:
            krbl['net_cost'][i] = krbl['flt_rate'][i] * krbl['Chg'][i]
            if krbl['FlightNo.'][i][:2] == 'QP':
                krbl['net_cost'][i] = max(krbl['flt_rate'][i] * krbl['Chg'][i], 1400) + max(krbl['Chg'][i] * 1, 300) + 300
            elif krbl['FlightNo.'][i][:2] == 'IX':
                krbl['net_cost'][i] = max(krbl['flt_rate'][i] * krbl['Chg'][i], 1500) + max(krbl['Chg'][i] * 2, 500)
            elif krbl['FlightNo.'][i][:2] == 'SG':
                krbl['net_cost'][i] = max(krbl['flt_rate'][i] * krbl['Chg'][i], 1600)
            elif krbl['FlightNo.'][i][:2] == '6E':
                krbl['net_cost'][i] = max(krbl['flt_rate'][i] * krbl['Chg'][i], 1500)
            elif krbl['FlightNo.'][i][:2] == 'AI':
                krbl['net_cost'][i] = max(krbl['flt_rate'][i] * krbl['Chg'][i], 1500) + max(krbl['Chg'][i] * 2, 500)
        except Exception as e:
            continue

    #flight cost, minimum charges comparison to be included
    krbl['net_cost']=''
    for i in range(len(krbl)):
        krbl['net_cost'][i]=krbl['flt_rate'][i]*krbl['Chg'][i]
        if krbl['FlightNo.'][i][:2]=='QP':
            krbl['net_cost'][i]=max(krbl['flt_rate'][i]*krbl['Chg'][i],1400)+max(krbl['Chg'][i]*1,300)+300
        elif krbl['FlightNo.'][i][:2]=='IX':
            krbl['net_cost'][i]=max(krbl['flt_rate'][i]*krbl['Chg'][i],1500)+max(krbl['Chg'][i]*2,500)
        elif krbl['FlightNo.'][i][:2]=='SG':
            krbl['net_cost'][i]=max(krbl['flt_rate'][i]*krbl['Chg'][i],1600)
        elif krbl['FlightNo.'][i][:2]=='6E':
            krbl['net_cost'][i]=max(krbl['flt_rate'][i]*krbl['Chg'][i],1500)
        elif krbl['FlightNo.'][i][:2]=='AI':
            krbl['net_cost'][i]=max(krbl['flt_rate'][i]*krbl['Chg'][i],1500)+max(krbl['Chg'][i]*2,500)

    krbl.head(20)

    # INDEX

    index=clh[(clh['Co-Loader']=='Index')]
    index=index.reset_index(drop=True)

    import pandas as pd
    index_rates=pd.read_excel(rates_file,sheet_name='index')
    index['flt_rate']=''
    index['FlightNo.']=index['FlightNo.'].astype(str)
    for i in range(len(index)):
        for j in range(len(index_rates)):
            if index['Origin'][i]+index['Destination'][i]+index['FlightNo.'][i]==index_rates['key'][j]:
                if index['Chg'][i]<300:
                    index['flt_rate'][i]=index_rates['+100'][j]
                elif index['Chg'][i]<500:
                    index['flt_rate'][i]=index_rates['+300'][j]
                else:
                    index['flt_rate'][i]=index_rates['+500'][j]
                break
            elif index['Origin'][i]+index['Destination'][i]+index['FlightNo.'][i][:2]+"9999"==index_rates['key'][j]:
                if pd.isnull(index_rates['start_time'][j]):
                    if index['Chg'][i]<300:
                        index['flt_rate'][i]=index_rates['+100'][j]
                    elif index['Chg'][i]<500:
                        index['flt_rate'][i]=index_rates['+300'][j]
                    else:
                        index['flt_rate'][i]=index_rates['+500'][j]
                break

    #Rates for flights with time constraints
    for j in range(len(index_rates)):
        for i in range(len(index)):
            if pd.isnull(index_rates['start_time'][j]) is False and index_rates['key'][j]==index['Origin'][i]+index['Destination'][i]+index['FlightNo.'][i][:2]+"9999":
                if index['ATD'][i]>=index_rates['start_time'][j] and index['ATD'][i]<=index_rates['end_time'][j]:
                    if index['Chg'][i]<300:
                        index['flt_rate'][i]=index_rates['+100'][j]
                    elif index['Chg'][i]<500:
                        index['flt_rate'][i]=index_rates['+300'][j]
                    else:
                        index['flt_rate'][i]=index_rates['+500'][j]

    index['flt_rate']=index['flt_rate'].apply(pd.to_numeric)

    #flight cost, minimum charges comparison to be included 
    index['net_cost'] = ''
    for i in range(len(index)):
        try:
            index['net_cost'][i] = index['flt_rate'][i] * index['Chg'][i]
            if index['FlightNo.'][i][:2] == 'QP':
                index['net_cost'][i] = max(index['flt_rate'][i] * index['Chg'][i], 1400)
            elif index['FlightNo.'][i][:2] == 'G8':
                index['net_cost'][i] = max(index['flt_rate'][i] * index['Chg'][i], 1500)
            elif index['FlightNo.'][i][:2] == 'SG':
                index['net_cost'][i] = max(index['flt_rate'][i] * index['Chg'][i], 1600)
            elif index['FlightNo.'][i][:2] == '6E':
                index['net_cost'][i] = max(index['flt_rate'][i] * index['Chg'][i], 1500)
            elif index['FlightNo.'][i][:2] == 'AI':
                index['net_cost'][i] = max(index['flt_rate'][i] * index['Chg'][i], 1500)
        except Exception as e:
            continue

    #AWB+DO
    for i in range(len(index)):
        try:
            index['net_cost'][i]=index['net_cost'][i]+500
        except Exception as e:
            continue



    # POBC 

    pobc=clh[(clh['Co-Loader']=='Patel On-Board Couriers Ltd')]
    pobc=pobc.reset_index(drop=True)

    #surface delivery

    delivery_cost={
        "lane":['IXA-IXA','SIL-SIL','VNS-VNS','VTZ-VTZ','IXC-IXC','RPR-RPR','IXR-IXR','PAT-PAT'],
        "rate":[900,1600,1200,1700,1250,1400,1100,2000]
    }
    surface=pd.DataFrame(delivery_cost)

    pobc['flt_rate']=''
    for i in range(len(pobc)):
        for j in range(len(surface)):
            if pobc['Airline'][i]=='Surface' and pobc['Lane'][i]==surface['lane'][j]:
                pobc['flt_rate'][i]=surface['rate'][j]


    #POBC_flt_rates
    pobc_rates_dedi=pd.read_excel(rates_file,sheet_name='pobc-dedicated')
    pobc_rates_consol=pd.read_excel(rates_file,sheet_name='pobc-consol')
    airline_tariff=pd.read_excel(rates_file,sheet_name='airline_tariff')

    pobc['FlightNo.']=pobc['FlightNo.'].astype(str)
    for i in range(len(pobc)):
        if pobc['Mode'][i]!='Consolidated':
            for j in range(len(pobc_rates_dedi)):
                if pobc['Origin'][i]+pobc['Destination'][i]+pobc['FlightNo.'][i]==pobc_rates_dedi['key'][j]:
                    if pobc['Chg'][i]>=1000:
                        pobc['flt_rate'][i]=pobc_rates_dedi['+1000'][j]
                    elif pobc['Chg'][i]>=500:
                        pobc['flt_rate'][i]=pobc_rates_dedi['+500'][j]
                    elif pobc['Chg'][i]>=300:
                        pobc['flt_rate'][i]=pobc_rates_dedi['+300'][j]
                    elif pobc['Chg'][i]>=100:
                        pobc['flt_rate'][i]=pobc_rates_dedi['+100'][j]
                    elif pobc['Chg'][i]>=45:
                        pobc['flt_rate'][i]=pobc_rates_dedi['+45'][j]
                    else:
                        for m in range(len(airline_tariff)):
                            if pobc['Origin'][i]+pobc['Destination'][i]==airline_tariff['key'][m] and pobc['FlightNo.'][i][:2]==airline_tariff['flight'][m]:
                                pobc['flt_rate'][i]=airline_tariff['n_rate'][m]
                    break

                elif pobc['Origin'][i]+pobc['Destination'][i]+pobc['FlightNo.'][i][:2]+"9999"==pobc_rates_dedi['key'][j]:
                    if pd.isnull(pobc_rates_dedi['start_time'][j]):
                        if pobc['Chg'][i]>=1000:
                            pobc['flt_rate'][i]=pobc_rates_dedi['+1000'][j]
                        elif pobc['Chg'][i]>=500:
                            pobc['flt_rate'][i]=pobc_rates_dedi['+500'][j]
                        elif pobc['Chg'][i]>=300:
                            pobc['flt_rate'][i]=pobc_rates_dedi['+300'][j]
                        elif pobc['Chg'][i]>=100:
                            pobc['flt_rate'][i]=pobc_rates_dedi['+100'][j]
                        elif pobc['Chg'][i]>=45:
                            pobc['flt_rate'][i]=pobc_rates_dedi['+45'][j]
                        else:
                            for m in range(len(airline_tariff)):
                                if pobc['Origin'][i]+pobc['Destination'][i]==airline_tariff['key'][m] and pobc['FlightNo.'][i][:2]==airline_tariff['flight'][m]:
                                    pobc['flt_rate'][i]=airline_tariff['n_rate'][m]

        else:
            for k in range(len(pobc_rates_consol)):
                if pobc['Origin'][i]+pobc['Destination'][i]+pobc['FlightNo.'][i]==pobc_rates_consol['key'][k]:
                    if pobc['Chg'][i]>=1000:
                        pobc['flt_rate'][i]=pobc_rates_consol['+1000'][k]
                    elif pobc['Chg'][i]>=500:
                        pobc['flt_rate'][i]=pobc_rates_consol['+500'][k]
                    elif pobc['Chg'][i]>=300:
                        pobc['flt_rate'][i]=pobc_rates_consol['+300'][k]
                    elif pobc['Chg'][i]>=100:
                        pobc['flt_rate'][i]=pobc_rates_consol['+100'][k]
                    elif pobc['Chg'][i]>=45:
                        pobc['flt_rate'][i]=pobc_rates_consol['+45'][k]
                    else:
                        for m in range(len(airline_tariff)):
                            if pobc['Origin'][i]+pobc['Destination'][i]==airline_tariff['key'][m] and pobc['FlightNo.'][i][:2]==airline_tariff['flight'][m]:
                                pobc['flt_rate'][i]=airline_tariff['n_rate'][m]
                    break

                elif pobc['Origin'][i]+pobc['Destination'][i]+pobc['FlightNo.'][i][:2]+"9999"==pobc_rates_consol['key'][k]:
                    if pd.isnull(pobc_rates_consol['start_time'][k]):
                        if pobc['Chg'][i]>=1000:
                            pobc['flt_rate'][i]=pobc_rates_consol['+1000'][k]
                        elif pobc['Chg'][i]>=500:
                            pobc['flt_rate'][i]=pobc_rates_consol['+500'][k]
                        elif pobc['Chg'][i]>=300:
                            pobc['flt_rate'][i]=pobc_rates_consol['+300'][k]
                        elif pobc['Chg'][i]>=100:
                            pobc['flt_rate'][i]=pobc_rates_consol['+100'][k]
                        elif pobc['Chg'][i]>=45:
                            pobc['flt_rate'][i]=pobc_rates_consol['+45'][k]
                        else:
                            for m in range(len(airline_tariff)):
                                if pobc['Origin'][i]+pobc['Destination'][i]==airline_tariff['key'][m] and pobc['FlightNo.'][i][:2]==airline_tariff['flight'][m]:
                                    pobc['flt_rate'][i]=airline_tariff['n_rate'][m]

    #Rates for flights with time constraint
    from datetime import datetime
    pobc.reset_index(drop=True, inplace=True)

    for j in range(len(pobc_rates_dedi)):
        for i in range(len(pobc)):
            if pobc['Mode'][i] != 'Consolidated':
                if pd.isnull(pobc_rates_dedi['start_time'][j]) is False and pobc_rates_dedi['key'][j] == pobc['Origin'][i] + pobc['Destination'][i] + pobc['FlightNo.'][i][:2] + "9999":
                    atd_time = datetime.strptime(pobc['ATD'][i], '%H:%M:%S').time()
                    if atd_time >= pobc_rates_dedi['start_time'][j] and atd_time <= pobc_rates_dedi['end_time'][j]:
                        if pobc['Chg'][i] >= 1000:
                            pobc['flt_rate'][i] = pobc_rates_dedi['+1000'][j]
                        elif pobc['Chg'][i] >= 500:
                            pobc['flt_rate'][i] = pobc_rates_dedi['+500'][j]
                        elif pobc['Chg'][i] >= 300:
                            pobc['flt_rate'][i] = pobc_rates_dedi['+300'][j]
                        elif pobc['Chg'][i] >= 100:
                            pobc['flt_rate'][i] = pobc_rates_dedi['+100'][j]
                        elif pobc['Chg'][i] >= 45:
                            pobc['flt_rate'][i] = pobc_rates_dedi['+45'][j]
                        else:
                            for m in range(len(airline_tariff)):
                                if pobc['Origin'][i] + pobc['Destination'][i] == airline_tariff['key'][m] and pobc['FlightNo.'][i][:2] == airline_tariff['flight'][m]:
                                    pobc['flt_rate'][i] = airline_tariff['n_rate'][m]

    for j in range(len(pobc_rates_consol)):
        for i in range(len(pobc)):
            if pobc['Mode'][i] == 'Consolidated':
                if pd.isnull(pobc_rates_consol['start_time'][j]) is False and pobc_rates_consol['key'][j] == pobc['Origin'][i] + pobc['Destination'][i] + pobc['FlightNo.'][i][:2] + "9999":
                    try:
                        if pobc['ATD'][i] >= pobc_rates_consol['start_time'][j] and pobc['ATD'][i] <= pobc_rates_consol['end_time'][j]:
                            if pobc['Chg'][i] >= 1000:
                                pobc['flt_rate'][i] = pobc_rates_consol['+1000'][j]
                            elif pobc['Chg'][i] >= 500:
                                pobc['flt_rate'][i] = pobc_rates_consol['+500'][j]
                            elif pobc['Chg'][i] >= 300:
                                pobc['flt_rate'][i] = pobc_rates_consol['+300'][j]
                            elif pobc['Chg'][i] >= 100:
                                pobc['flt_rate'][i] = pobc_rates_consol['+100'][j]
                            elif pobc['Chg'][i] >= 45:
                                pobc['flt_rate'][i] = pobc_rates_consol['+45'][j]
                            else:
                                for m in range(len(airline_tariff)):
                                    if pobc['Origin'][i] + pobc['Destination'][i] == airline_tariff['key'][m] and pobc['FlightNo.'][i][:2] == airline_tariff['flight'][m]:
                                        pobc['flt_rate'][i] = airline_tariff['n_rate'][m]
                    except KeyError:
                        pass

    pobc['flt_cost']=''
    pobc['flt_rate']=pobc['flt_rate'].apply(pd.to_numeric)
    flag=0
    for i in range(len(pobc)):
        try:
            if pobc['FlightNo.'][i][:2] == '6E':
                pobc['flt_cost'][i] = max(pobc['flt_rate'][i] * pobc['Chg'][i], 1500)
            elif pobc['FlightNo.'][i][:2] in ['AI', '9I']:
                pobc['flt_cost'][i] = max(pobc['flt_rate'][i] * pobc['Chg'][i], 1500)
            elif pobc['FlightNo.'][i][:2] == 'G8':
                pobc['flt_cost'][i] = max(pobc['flt_rate'][i] * pobc['Chg'][i], 1350)
            elif pobc['FlightNo.'][i][:2] == 'I5':
                pobc['flt_cost'][i] = max(pobc['flt_rate'][i] * pobc['Chg'][i], 900)
            elif pobc['FlightNo.'][i][:2] == 'SG':
                pobc['flt_cost'][i] = max(pobc['flt_rate'][i] * pobc['Chg'][i], 1200)
            elif pobc['FlightNo.'][i][:2] == 'UK':
                pobc['flt_cost'][i] = max(pobc['flt_rate'][i] * pobc['Chg'][i], 1400)
            elif pobc['FlightNo.'][i][:2] == 'QP':
                pobc['flt_cost'][i] = max(pobc['flt_rate'][i] * pobc['Chg'][i], 1400)
            elif pobc['Airline'][i] == 'Surface':
                pobc['flt_cost'][i] = pobc['flt_rate'][i]
        except Exception as e:
            flag=1
            continue
    print(flag)




    #CUT charges for <45 kg tenderments
    cut=pd.read_excel(rates_file,sheet_name='indigo-misc')
    pobc['cut_ob']=0
    pobc['cut_ib']=0

    #CUT_OUTBOUND_COST
    for i in range(len(pobc)):
        if pobc['Mode'][i]!='Delivery':
            if pobc['Chg'][i]<45:
                outbound=0
                for j in range(len(cut)):
                    if cut['station'][j]==pobc['Origin'][i]:
                        outbound=max(pobc['Chg'][i]*cut['OB_rate'][j],cut['OB_min'][j])+outbound
                pobc['cut_ob'][i]=outbound

    #CUT_INBOUND_COST
                inbound=0
                for j in range(len(cut)):
                    if cut['station'][j]==pobc['Destination'][i]:
                        inbound=max(pobc['Chg'][i]*cut['IB_rate'][j],cut['IB_min'][j])+inbound
                pobc['cut_ib'][i]=inbound

    pobc['awb+do']=''
    for i in range(len(pobc)):
        if pobc['Mode'][i]=='Delivery':
            pobc['awb+do'][i]=0
        elif pobc['Mode'][i]=='Consolidated':
            pobc['awb+do'][i]=150
        else:
            if pobc['FlightNo.'][i][:2]=='6E':
                pobc['awb+do'][i]=650
            elif pobc['FlightNo.'][i][:2]=='I5':
                pobc['awb+do'][i]=700+300
            elif pobc['FlightNo.'][i][:2]=='UK':
                pobc['awb+do'][i]=150
            elif pobc['FlightNo.'][i][:2]=='SG':
                pobc['awb+do'][i]=450
            elif pobc['FlightNo.'][i][:2]=='AI' or pobc['FlightNo.'][i][:2]=='9I':
                pobc['awb+do'][i]=max(3*pobc['Chg'][i],500)+125
            elif pobc['FlightNo.'][i][:2]=='QP':
                pobc['awb+do'][i]=300+max(1*pobc['Chg'][i],300)

    pobc['flt_cost']=pobc['flt_cost'].apply(pd.to_numeric)
    pobc['awb+do']=pobc['awb+do'].apply(pd.to_numeric)
    flag=1
    pobc['net_cost']=''
    for i in range(len(pobc)):
        try:
            pobc['net_cost'][i]=pobc['flt_cost'][i]+pobc['cut_ob'][i]+pobc['cut_ib'][i]+pobc['awb+do'][i]
        except Exception as e:
            flag=0
            print(flag)
            print(f"Error occurred for row {i}: {e}")
            continue


    pobc.head(5)

    # SURYA CARGO

    surya=clh[(clh['Co-Loader']=='Surya Cargo Forwarders Pvt Ltd')]
    surya=surya.reset_index(drop=True)
    surya.head(5)

    surya_rates=pd.read_excel(rates_file,sheet_name='surya')
    surya['flt_rate']=''
    surya['FlightNo.']=surya['FlightNo.'].astype(str)

    flt_rates_assigned = []

    for i in range(len(surya)):
        for j in range(len(surya_rates)):
            if surya['Origin'][i]+surya['Destination'][i]+surya['FlightNo.'][i]==surya_rates['key'][j]:
                if surya['Chg'][i]<100:
                    surya['flt_rate'][i]=surya_rates['+45'][j]
                    flt_rates_assigned.append(surya_rates['+45'][j])
                elif surya['Chg'][i]<300:
                    surya['flt_rate'][i]=surya_rates['+100'][j]
                    flt_rates_assigned.append(surya_rates['+100'][j])
                elif surya['Chg'][i]<500:
                    surya['flt_rate'][i]=surya_rates['+300'][j]
                    flt_rates_assigned.append(surya_rates['+300'][j])
                else:
                    surya['flt_rate'][i]=surya_rates['+500'][j]
                    flt_rates_assigned.append(surya_rates['+500'][j])
                print(f"flt_rate assigned for index {i}: {surya['flt_rate'][i]}")
                break
            elif surya['Origin'][i]+surya['Destination'][i]+surya['FlightNo.'][i][:2]+"9999"==surya_rates['key'][j]:
                if pd.isnull(surya_rates['start_time'][j]):
                    if surya['Chg'][i]<100:
                        surya['flt_rate'][i]=surya_rates['+45'][j]
                        flt_rates_assigned.append(surya_rates['+45'][j])
                    elif surya['Chg'][i]<300:
                        surya['flt_rate'][i]=surya_rates['+100'][j]
                        flt_rates_assigned.append(surya_rates['+100'][j])
                    elif surya['Chg'][i]<500:
                        surya['flt_rate'][i]=surya_rates['+300'][j]
                        flt_rates_assigned.append(surya_rates['+300'][j])
                    else:
                        surya['flt_rate'][i]=surya_rates['+500'][j]
                        flt_rates_assigned.append(surya_rates['+500'][j])
                    print(f"flt_rate assigned for index {i}: {surya['flt_rate'][i]}")
                break

    # Rates for flights with time constraints
    for j in range(len(surya_rates)):
        for i in range(len(surya)):
            if pd.isnull(surya_rates['start_time'][j]) is False and surya_rates['key'][j]==surya['Origin'][i]+surya['Destination'][i]+surya['FlightNo.'][i][:2]+"9999":
                if surya['ATD'][i]>=surya_rates['start_time'][j] and index['ATD'][i]<=index_rates['end_time'][j]:
                    if surya['Chg'][i]<100:
                        surya['flt_rate'][i]=surya_rates['+45'][j]
                        flt_rates_assigned.append(surya_rates['+45'][j])
                    elif surya['Chg'][i]<300:
                        surya['flt_rate'][i]=surya_rates['+100'][j]
                        flt_rates_assigned.append(surya_rates['+100'][j])
                    elif surya['Chg'][i]<500:
                        surya['flt_rate'][i]=surya_rates['+300'][j]
                        flt_rates_assigned.append(surya_rates['+300'][j])
                    else:
                        surya['flt_rate'][i]=surya_rates['+500'][j]
                        flt_rates_assigned.append(surya_rates['+500'][j])


    surya['flt_cost'] = ''

    missing_flt_rates = []
    flag = 0

    for i in range(len(surya)):
        try:
            flt_rate = float(surya['flt_rate'][i])
            chg = float(surya['Chg'][i])
            surya['flt_cost'][i] = max(flt_rate * chg, 1500)
        except Exception as e:
            flag = 1
            continue
    print(flag)

    surya['handling']=1*surya['Chg']
    surya['AWB+DO']=950

    #misc charges to be added for <100kg AWBs
    cut=pd.read_excel(rates_file,sheet_name='indigo-misc')
    surya['cut_ob']=0
    surya['cut_ib']=0

    #CUT_OUTBOUND_COST
    for i in range(len(surya)):
        if surya['Chg'][i]<100:
            outbound=0
            for j in range(len(cut)):
                if cut['station'][j]==surya['Origin'][i]:
                    outbound=max(surya['Chg'][i]*cut['OB_rate'][j],cut['OB_min'][j])+outbound
            surya['cut_ob'][i]=outbound

    #CUT_INBOUND_COST
            inbound=0
            for j in range(len(cut)):
                if cut['station'][j]==surya['Destination'][i]:
                    inbound=max(surya['Chg'][i]*cut['IB_rate'][j],cut['IB_min'][j])+inbound
            surya['cut_ib'][i]=inbound


    surya['net_cost'] = ''

    for i in range(len(surya)):
        try:
            surya['net_cost'][i] = surya['flt_cost'][i] + surya['handling'][i] + surya['AWB+DO'][i] + surya['cut_ob'][i] + surya['cut_ib'][i]
        except Exception as e:
            #print(f"Error occurred for row {i}: {e}")
            continue  
    with pd.ExcelWriter(output_file) as writer:
        surya.to_excel(writer,sheet_name='surya',index=False)


    # BHAGWATI AIR

    '''
    bhagwatiair=clh[(clh['Co-Loader']=='Bhagwati Air Express')]
    bhagwatiair=bhagwatiair.reset_index(drop=True)
    '''

    '''
    bhagwati_rates=pd.read_excel(rates_file,sheet_name='bhagwatiair')
    bhagwatiair['flt_rate']=''
    bhagwatiair['FlightNo.']=bhagwatiair['FlightNo.'].astype(str)
    for i in range(len(bhagwatiair)):
        for j in range(len(bhagwati_rates)):
            if bhagwatiair['Origin'][i]+bhagwatiair['Destination'][i]+bhagwatiair['FlightNo.'][i]==bhagwati_rates['key'][j]:
                if bhagwatiair['Mode'][i]=='Express':
                    bhagwatiair['flt_rate'][i]=bhagwati_rates['+100_rapid'][j]
                else:
                    bhagwatiair['flt_rate'][i]=bhagwati_rates['+100'][j]
                break
            elif bhagwatiair['Origin'][i]+bhagwatiair['Destination'][i]+bhagwatiair['FlightNo.'][i][:2]+"9999"==bhagwati_rates['key'][j]:
                if bhagwatiair['Mode'][i]=='Express':
                    bhagwatiair['flt_rate'][i]=bhagwati_rates['+100_rapid'][j]
                else:
                    bhagwatiair['flt_rate'][i]=bhagwati_rates['rate'][j]
    #bhagwatiair['flt_rate']=bhagwatiair['flt_rate'].apply(pd.to_numeric)
    '''

    #bhagwatiair['flt_cost']=bhagwatiair['flt_rate']*bhagwatiair['Chg']
    #bhagwatiair['AWB']=''
    #for i in range(len(bhagwatiair)):
    #    bhagwatiair['AWB'][i]=max(1*bhagwatiair['Chg'][i],500)

    #bhagwatiair['net_cost']=bhagwatiair['flt_cost']+bhagwatiair['AWB']

    #bhagwatiair.head(5)

    # Allied Express

    allied=clh[(clh['Co-Loader']=='Allied Express')]
    allied=allied.reset_index(drop=True)

    allied_rates=pd.read_excel(rates_file,sheet_name='allied')
    allied['flt_rate']=''
    allied['FlightNo.']=allied['FlightNo.'].astype(str)
    for i in range(len(allied)):
        for j in range(len(allied_rates)):
            if allied['Origin'][i]+allied['Destination'][i]+allied['FlightNo.'][i]==allied_rates['key'][j]:
                if allied['Chg'][i]<300:
                    allied['flt_rate'][i]=allied_rates['+100'][j]
                elif allied['Chg'][i]<500:
                    allied['flt_rate'][i]=allied_rates['+300'][j]
                elif allied['Chg'][i]<1000:
                    allied['flt_rate'][i]=allied_rates['+500'][j]
                else:
                    allied['flt_rate'][i]=allied_rates['+1000'][j]
                break
            elif allied['Origin'][i]+allied['Destination'][i]+allied['FlightNo.'][i][:2]+"9999"==allied_rates['key'][j]:
                if allied['Chg'][i]<300:
                    allied['flt_rate'][i]=allied_rates['+100'][j]
                elif allied['Chg'][i]<500:
                    allied['flt_rate'][i]=allied_rates['+300'][j]
                elif allied['Chg'][i]<1000:
                    allied['flt_rate'][i]=allied_rates['+500'][j]
                else:
                    allied['flt_rate'][i]=allied_rates['+1000'][j]

    allied['flt_rate']=allied['flt_rate'].apply(pd.to_numeric)

    allied['flt_cost']=''

    for i in range(len(allied)):
        try:
            flt_rate = float(allied['flt_rate'][i])
            chg = float(allied['Chg'][i])
            allied['flt_cost'][i] = max(flt_rate * chg, 1500)
        except Exception as e:
            flag = 1
            continue

    print(flag)


    #AWB
    allied['awb+do']=''
    for i in range(len(allied)):
        if allied['FlightNo.'][i][:2]=='6E':
            allied['awb+do'][i]=150+max(8*allied['Chg'][i],800)
        elif allied['FlightNo.'][i][:2]=='I5':
            allied['awb+do'][i]=700+300
        elif allied['FlightNo.'][i][:2]=='UK':
            allied['awb+do'][i]=200+max(3*allied['Chg'][i],300)
        elif allied['FlightNo.'][i][:2]=='G8':
            allied['awb+do'][i]=300+max(1*allied['Chg'][i],550)
        elif allied['FlightNo.'][i][:2]=='AI' or allied['FlightNo.'][i][:2]=='9I':
            allied['awb+do'][i]=max(3*allied['Chg'][i],500)+125
        elif allied['FlightNo.'][i][:2]=='QP':
            allied['awb+do'][i]=300+max(1*allied['Chg'][i],300)

    allied['net_cost']=''
    for i in range(len(allied)):
        try:
            allied['net_cost'][i]=allied['flt_cost'][i]+allied['awb+do'][i]
        except Exception as e:
            print(f"Error occurred for row {i}: {e}")
            continue
    with pd.ExcelWriter(output_file) as writer:
        allied.to_excel(writer,sheet_name='nayak',index=False)



    allied.head(5)

    # LABRIYNTH LOGISTICS PVT LTD

    labyrinth=clh[(clh['Co-Loader']=='Labyrinth Logistics Pvt Ltd')]
    labyrinth=labyrinth.reset_index(drop=True)

    labyrinth['rate']=''
    for i in range(len(labyrinth)):
        if labyrinth['Lane'][i]=="MAA-IXZ":
            labyrinth['rate'][i]=20.625
        elif labyrinth['Lane'][i]=="IXZ-MAA":
            labyrinth['rate'][i]=23.25
        elif labyrinth['Lane'][i]=="COK-AGX" or labyrinth['Lane'][i]=="AGX-COK":
            if labyrinth['Chg'][i]<100:
                labyrinth['rate'][i]=2000/labyrinth['Chg'][i]
            elif labyrinth['Chg'][i]<250:
                labyrinth['rate'][i]=35
            elif labyrinth['Chg'][i]<500:
                labyrinth['rate'][i]=25
            else:
                labyrinth['rate'][i]=20

    labyrinth['rate']=labyrinth['rate'].apply(pd.to_numeric)
    labyrinth['net_cost']=labyrinth['rate']*labyrinth['Chg']

    failed_df.head(25)

    # Compilation

    accruals=pd.concat([indigo.iloc[:,:17],spicejet.iloc[:,:17],airasia.iloc[:,:17],vistara.iloc[:,:17],krbl.iloc[:,:17],pobc.iloc[:,:17],surya.iloc[:,:17],allied.iloc[:,:17],labyrinth.iloc[:,:17],index.iloc[:,:17]])
    accruals=accruals.reset_index(drop=True)

    net_cost=pd.concat([indigo['net_cost'],spicejet['net_cost'],airasia['net_cost'],vistara['net_cost'],krbl['net_cost'],pobc['net_cost'],surya['net_cost'],allied['net_cost'],labyrinth['net_cost'],index['net_cost']])
    net_cost=net_cost.reset_index(drop=True)

    accruals['net_cost']=''
    for i in range(len(accruals)):
        accruals['net_cost'][i]=net_cost[i]

    accruals['Week']=''
    accruals['Date'] = pd.to_datetime(accruals['Date'])
    accruals['Week']=accruals['Date'].apply(lambda x: x.strftime('%U'))
    specified_origins = ["AMD", "BLR", "BOM", "CCU", "CJB", "COK", "DEL", "HYD", "IXC", "IXL","IXR", "IXZ", "LKO", "MAA", "PAT", "PNQ", "RPR", "VTZ"]


    # BATS

    bats=clh[(clh['Origin']=='BOM')]
    bats=bats.reset_index(drop=True)

    bats['rate']=2.1
    bats['bill_amt']=2.1*bats['Chg']
    bats['misc']=(264279/bats['Chg'].sum())*bats['Chg']

    bats['net_cost']=bats['bill_amt']+bats['misc']
    def check_origin(origin):
        return origin in specified_origins
    
    accruals['Origin_Check'] = accruals['Origin'].apply(check_origin)
    accruals['CpKG'] = '-'


    non_zero_chg_indices = accruals['Chg'] != 0
    accruals.loc[non_zero_chg_indices, 'CpKG'] = accruals.loc[non_zero_chg_indices, 'net_cost'] / accruals.loc[non_zero_chg_indices, 'Chg']
    accruals['ATD'] = pd.to_datetime(accruals['ATD'], errors='coerce')

    # Define the calculate_MAE function
    def calculate_MAE(row):
        if pd.isnull(row['ATD']):
            return 'Missing'  
        atd_time = row['ATD'].time()
        if (atd_time >= pd.Timestamp('02:00:00').time() and atd_time <= pd.Timestamp('09:00:00').time()) or row['Airline']=='Amazon Air':
            return 'M'
        elif (atd_time > pd.Timestamp('09:00:00').time() and atd_time <= pd.Timestamp('17:00:00').time()):
            return 'A'
        else:
            return 'E'


    accruals['M/A/E'] = accruals.apply(calculate_MAE, axis=1)
    accruals['Date'] = pd.to_datetime(accruals['Date']).dt.strftime('%m/%d/%Y')
    accruals['ATD'] = pd.to_datetime(accruals['ATD']).dt.strftime('%H:%M:%S')

    # BLUEDART

    bluedart=clh[(clh['Co-Loader']=='BlueDart')]
    bluedart=bluedart.reset_index(drop=True)

    bluedart_rates=pd.read_excel(rates_file,sheet_name='bluedart')
    bluedart['_rate']=''

    bluedart['flt_rate']=''
    for i in range(len(bluedart)):
        for j in range(len(bluedart_rates)):
            if bluedart['Lane'][i]==bluedart_rates['lane'][j]:
                bluedart['flt_rate'][i]=bluedart_rates['rate'][j]

    bluedart['flt_cost']=bluedart['flt_rate']*bluedart['Chg']

    # Summary

    data={
        "airlines":['indigo','spicejet','airasia','vistara','krbl','pobc','surya','allied','labyrinth','index','belly+sea+freight/express','belly+sea','belly']
    }
    summary=pd.DataFrame(data)

    summary['cpkg']=''
    summary['cpkg'][0]=indigo['net_cost'].sum()/indigo['Chg'].sum()
    summary['cpkg'][1]=spicejet['net_cost'].sum()/spicejet['Chg'].sum()
    summary['cpkg'][2]=airasia['net_cost'].sum()/airasia['Chg'].sum()
    summary['cpkg'][3]=vistara['net_cost'].sum()/vistara['Chg'].sum()
    summary['cpkg'][4]=krbl['net_cost'].sum()/krbl['Chg'].sum()
    summary['cpkg'][5]=pobc['net_cost'].sum()/pobc['Chg'].sum()
    summary['cpkg'][6]=surya['net_cost'].sum()/surya['Chg'].sum()
    #summary['cpkg'][7]=bhagwatiair['net_cost'].sum()/bhagwatiair['Chg'].sum()
    summary['cpkg'][8]=allied['net_cost'].sum()/allied['Chg'].sum()
    summary['cpkg'][9]=index['net_cost'].sum()/index['Chg'].sum()
    summary['cpkg'][10]=labyrinth['net_cost'].sum()/labyrinth['Chg'].sum()

    #BELLY + SEA + FREIGHT/EXPRESS
    summary['cpkg'][10]=accruals['net_cost'].sum()/accruals['Chg'].sum()

    #BELLY + SEA CPKG
    summary['cpkg'][11]=sum(accruals[accruals['Mode']!='Express']['net_cost'])/sum(accruals[accruals['Mode']!='Express']['Chg'])

    #BELLY CPKG
    summary['cpkg'][12]=sum(accruals[(accruals['Mode']!='Express') & (accruals['Airline']!='Sea')]['net_cost'])/sum(accruals[(accruals['Mode']!='Express') & (accruals['Airline']!='Sea')]['Chg'])

    #prime/GCR share
    summary['category']=''
    summary['category'][0]='prime'
    summary['category'][1]='gcr'
    summary['share']=''
    summary['share'][0]=(sum(indigo[indigo['flt_type']=='prime']['Chg'])/indigo['Chg'].sum())*100
    summary['share'][1]=(sum(indigo[indigo['flt_type']=='non-prime']['Chg'])/indigo['Chg'].sum())*100

    weekn=max(accruals['Week'])
    week=accruals[accruals['Week']==weekn]

    coloaders = {
        "coloaders": ['indigo', 'spicejet', 'krbl', 'pobc', 'surya', 'bhagwati', 'nayak', 'broekman', 'index', 'belly+sea+freight/express', 'belly+sea', 'belly']
    }
    week_summary = pd.DataFrame(coloaders)
    week_summary['cpkg'] = ''
    try:
        week_summary['cpkg'][0] = sum(week[week['Co-Loader'] == 'IndiGo Air']['net_cost']) / sum(week[week['Co-Loader'] == 'IndiGo Air']['Chg'])
    except:
        week_summary['cpkg'][0] = 0
    try:
        week_summary['cpkg'][1] = sum(week[week['Co-Loader'] == 'SpiceJet']['net_cost']) / sum(week[week['Co-Loader'] == 'SpiceJet']['Chg'])
    except:
        week_summary['cpkg'][1] = 0
    try:
        week_summary['cpkg'][2] = sum(week[week['Co-Loader'] == 'Kreative Rainbowbliss Logistics']['net_cost']) / sum(week[week['Co-Loader'] == 'Kreative Rainbowbliss Logistics']['Chg'])
    except:
        week_summary['cpkg'][2] = 0
    #try:
    #    week_summary['cpkg'][3] = sum(week[week['Co-Loader'] == 'Bhagwati Air Express']['net_cost']) / sum(week[week['Co-Loader'] == 'Bhagwati Air Express']['Chg'])
    #except:
    #    week_summary['cpkg'][3]
    try:
        week_summary['cpkg'][4] = sum(week[week['Co-Loader'] == 'Patel On-Board Couriers Ltd']['net_cost']) / sum(week[week['Co-Loader'] == 'Patel On-Board Couriers Ltd']['Chg'])
    except:
        week_summary['cpkg'][4] = 0
    try:
        week_summary['cpkg'][5] = sum(week[week['Co-Loader'] == 'Surya Cargo Forwarders Pvt Ltd']['net_cost']) / sum(week[week['Co-Loader'] == 'Surya Cargo Forwarders Pvt Ltd']['Chg'])
    except:
        week_summary['cpkg'][5] = 0
    try:
        week_summary['cpkg'][6] = sum(week[week['Co-Loader'] == 'Allied Express']['net_cost']) / sum(week[week['Co-Loader'] == 'Allied Express']['Chg'])
    except:
        week_summary['cpkg'][6] = 0
    try:
        week_summary['cpkg'][7] = sum(week[week['Co-Loader'] == 'Labyrinth Logistics Pvt Ltd']['net_cost']) / sum(week[week['Co-Loader'] == 'Labyrinth Logistics Pvt Ltd']['Chg'])
    except:
        week_summary['cpkg'][7] = 0
    try:
        week_summary['cpkg'][8] = sum(week[week['Co-Loader'] == 'index']['net_cost']) / sum(week[week['Co-Loader'] == 'index']['Chg'])
    except:
        week_summary['cpkg'][8] = 0

    #BELLY + SEA + FREIGHT/EXPRESS
    week_summary['cpkg'][8]=week['net_cost'].sum()/week['Chg'].sum()

    #BELLY + SEA CPKG
    week_summary['cpkg'][9]=sum(week[week['Mode']!='Express']['net_cost'])/sum(week[week['Mode']!='Express']['Chg'])

    #BELLY CPKG
    week_summary['cpkg'][10]=sum(week[(week['Mode']!='Express') & (week['Airline']!='Sea')]['net_cost'])/sum(week[(week['Mode']!='Express') & (week['Airline']!='Sea')]['Chg'])

    indigo['Date'] = pd.to_datetime(indigo['Date'])
    indigo['Week']=''
    for i in range(len(indigo)):
        indigo['Week'][i]=indigo['Date'][i].week

    week_indigo=indigo[indigo['Week']==weekn]

    week_summary['category']=''
    week_summary['category'][0]='prime'
    week_summary['category'][1]='gcr'
    week_summary['share']=''
    week_summary['share'][0]=(sum(week_indigo[week_indigo['flt_type']=='prime']['Chg'])/week_indigo['Chg'].sum())*100
    week_summary['share'][1]=(sum(week_indigo[week_indigo['flt_type']=='non-prime']['Chg'])/week_indigo['Chg'].sum())*100

    #EXPORT TO EXCEL
    with pd.ExcelWriter(output_file) as writer:
        indigo.to_excel(writer,sheet_name='indigo',index=False)
        spicejet.to_excel(writer,sheet_name='spicejet',index=False)
        airasia.to_excel(writer,sheet_name='airasia',index=False)
        vistara.to_excel(writer,sheet_name='vistara',index=False)
        krbl.to_excel(writer,sheet_name='krbl',index=False)
        pobc.to_excel(writer,sheet_name='pobc',index=False)
        surya.to_excel(writer,sheet_name='surya',index=False)
        #bhagwatiair.to_excel(writer,sheet_name='bhagwati',index=False)
        allied.to_excel(writer,sheet_name='nayak',index=False)
        labyrinth.to_excel(writer,sheet_name='labyrinth',index=False)
        index.to_excel(writer,sheet_name='Index',index=False)
        accruals.to_excel(writer,sheet_name='accruals',index=False)
        bats.to_excel(writer,sheet_name='bats',index=False)
        summary.to_excel(writer,sheet_name='summary',startrow=1)
        sheet=writer.sheets['summary']
        sheet.write(0,0,"Month to Date")
        week_summary.to_excel(writer,sheet_name='summary',startrow=18)
        sheet.write(17,0,"Previous Week")
    writer.close()

    #Quickjet Cost

    import numpy as np
    import pandas as pd
    import calendar
    from datetime import datetime as dt
    from datetime import time, timedelta
    qj_mis=pd.read_excel(qj_mis_file,sheet_name='Sales')
    qj_mis.drop_duplicates(subset=['Awb No'],inplace=True)
    qj_mis.dropna(inplace=True)
    qj_mis=qj_mis[qj_mis['Shipper - Account name']!='Quikjet Comat']
    qj_mis['lane']=qj_mis['Origin - Airport code']+"-"+qj_mis['Destination - Airport code']
    leg={
        "lane":['DEL-BOM','DEL-BLR','DEL-HYD','BLR-DEL','HYD-DEL','BOM-DEL','HYD-BOM','BLR-HYD'],
        "leg":['DEL-BOM','DEL-BLR','DEL-BLR','BLR-HYD','HYD-DEL','BOM-DEL','HYD-DEL','BLR-HYD']
        }
    leg_mapping=pd.DataFrame(leg)

    qj_mis=qj_mis.merge(leg_mapping[['lane','leg']],on='lane',how='left')

    qj_mis = qj_mis.dropna()
    pnq_vols=pd.read_excel(file_path_1,'Sheet1')

    qj_mis=qj_mis.merge(pnq_vols[['Awb No','lane_updated']],on='Awb No',how='left')

    qj_mis['lane_updated']=qj_mis.apply(lambda row:row['lane'] if pd.isnull(row['lane_updated']) else row['lane_updated'],axis=1)


    AAZ=['DEL-BOM','BOM-DEL']
    BAZ=['HYD-DEL','DEL-BLR','BLR-HYD']

    qj_mis['aircraft']=qj_mis.apply(lambda row:'AAZ' if row['leg'] in AAZ else 'BAZ',axis=1)



    "QuikJet costs"

    fixed_cost=31880059
    bh_cost=88992
    cycle_cost=113511
    cut_ob={
         "Origin - Airport code":['HYD','DEL','BLR','BOM'],
         "rate_ob":[6.28,6.46,5.24,5.54]
         }
    cut_ob=pd.DataFrame(cut_ob)

    cut_ib={
            "Destination - Airport code":['HYD','DEL','BLR','BOM'],
            "rate_ib":[4.11,3,2.96,2.92]
            }
    cut_ib=pd.DataFrame(cut_ib)

    "ADDING FIXED COSTS"

    now=dt.now()-timedelta(days=2)
    days=calendar.monthrange(now.year,now.month)[1]


    qj_mis['Flight Date Local']=pd.to_datetime(qj_mis['Flight Date Local'],format='%d %b %Y')

    '''
    for i in range(len(qj_mis)):
        if qj_mis['lane_updated'][i]=='HYD-DEL':
            qj_mis['Flight Date Local'][i]=qj_mis['Flight Date Local'][i]+pd.Timedelta(days=1)
    '''
    otp=pd.read_excel(file_path_2,'Sheet1')
    '''
    for i in range(len(otp)):
        if otp['Sector'][i]=='HYD-DEL':
            otp['Day'][i]=otp['Day'][i]+pd.Timedelta(days=1)
    '''
    otp['key']=otp['Day'].dt.strftime('%Y-%m-%d') + otp['Sector']


    cost=qj_mis.groupby(['Flight Date Local','leg'])['Total Chargeable Weight'].sum()
    cost=pd.DataFrame(cost)
    cost.reset_index(inplace=True)
    cost['aircraft']=cost.apply(lambda row:'AAZ' if row['leg'] in AAZ else 'BAZ',axis=1)
    cost['key']=cost['Flight Date Local'].dt.strftime('%Y-%m-%d') + cost['leg']
    cost.head()
    cost=cost.merge(otp[['key','BT']],on='key',how='left')
    cost.dropna(inplace=True)
    cost.reset_index(inplace=True)
    cost['BT']=pd.to_datetime(cost['BT'],format='%H:%M:%S',errors='coerce').dt.time
    cost['block_hours']=cost['BT'].apply(lambda x:x.hour + x.minute/60)
    cost['BH_cost']=bh_cost*cost['block_hours']
    cost['BH_cost_perkg']=cost['BH_cost']/cost['Total Chargeable Weight']
    count=cost.groupby(['Flight Date Local','aircraft'])['aircraft'].count()
    count=pd.DataFrame(count)
    count.rename(columns={'aircraft':'count'},inplace=True)
    count.reset_index(inplace=True)

    cost['fixed_cost']=''
    for i in range(len(cost)):
        for j in range(len(count)):
            if cost['Flight Date Local'][i]==count['Flight Date Local'][j] and cost['aircraft'][i]==count['aircraft'][j]:
                cost['fixed_cost'][i]=(fixed_cost/days)/count['count'][j]

    cost['fixed_cost_perkg']=cost['fixed_cost']/cost['Total Chargeable Weight']
    cost['cycle_cost']=cycle_cost
    cost['cycle_cost_perkg']=cost['cycle_cost']/cost['Total Chargeable Weight']

    parking={
        "leg":["HYD-DEL","DEL-BLR","BLR-HYD","DEL-BOM","BOM-DEL"],
        "parking":[49093,37357,36157,46377,68828]
        }
    parking=pd.DataFrame(parking)

    handling={
        "leg":["HYD-DEL","DEL-BLR","BLR-HYD","DEL-BOM","BOM-DEL"],
        "handling":[39313,40453,40453,40453,39313]
        }
    handling=pd.DataFrame(handling)

    cost=cost.merge(parking[['leg','parking']],on='leg',how='left')
    cost['parking_perkg']=cost['parking']/cost['Total Chargeable Weight']

    cost=cost.merge(handling[['leg','handling']],on='leg',how='left')
    cost['handling_perkg']=cost['handling']/cost['Total Chargeable Weight']

    qj_mis['key']=qj_mis['Flight Date Local'].dt.strftime('%Y-%m-%d')+qj_mis['leg']

    qj_mis=qj_mis.merge(otp[['key','Flt.No']],on="key",how='left')

    qj_mis=qj_mis.merge(cost[['key','BH_cost_perkg','fixed_cost_perkg','cycle_cost_perkg','parking_perkg','handling_perkg']],on='key',how='left')

    qj_mis['BH_cost']=qj_mis['Total Chargeable Weight']*qj_mis['BH_cost_perkg']
    qj_mis.drop(columns=['BH_cost_perkg'],inplace=True)

    qj_mis['fixed_cost']=qj_mis['Total Chargeable Weight']*qj_mis['fixed_cost_perkg']
    qj_mis.drop(columns=['fixed_cost_perkg'],inplace=True)

    qj_mis['cycle_cost']=qj_mis['Total Chargeable Weight']*qj_mis['cycle_cost_perkg']
    qj_mis.drop(columns=['cycle_cost_perkg'],inplace=True)

    qj_mis['parking_cost']=qj_mis['Total Chargeable Weight']*qj_mis['parking_perkg']
    qj_mis.drop(columns=['parking_perkg'],inplace=True)

    qj_mis['handling_cost']=qj_mis['Total Chargeable Weight']*qj_mis['handling_perkg']
    qj_mis.drop(columns=['handling_perkg'],inplace=True)

    qj_mis=qj_mis.merge(cut_ob[['Origin - Airport code','rate_ob']],on="Origin - Airport code",how='left')
    qj_mis['cut_ob']=qj_mis['Total Chargeable Weight']*qj_mis['rate_ob']
    qj_mis.drop(columns=['rate_ob'],inplace=True)

    qj_mis=qj_mis.merge(cut_ib[['Destination - Airport code','rate_ib']],on="Destination - Airport code",how='left')
    qj_mis['cut_ib']=qj_mis['Total Chargeable Weight']*qj_mis['rate_ib']
    qj_mis.drop(columns=['rate_ib'],inplace=True)

    qj_mis['cut_cost']=qj_mis['cut_ob']+qj_mis['cut_ib']

    cut=qj_mis.groupby(['Flight Date Local','leg'])['cut_cost'].sum()
    cut=pd.DataFrame(cut)
    cut.reset_index(inplace=True)
    cut['key']=cut['Flight Date Local'].dt.strftime('%Y-%m-%d') + cut['leg']

    cost=cost.merge(cut[['key','cut_cost']],on='key',how='left')

    "FUEL COST CALCULATION"
    fuel_cost={
        "origin":["DEL","HYD","BLR","BOM"],
        "rate":[82.26,69.37,77.69,76.88]
        }
    fuel_cost=pd.DataFrame(fuel_cost)

    otp['origin']=otp['Sector'].str[:3]

    otp=otp.merge(fuel_cost[['origin','rate']],on='origin',how='left')

    otp['fuel_upliftment_cost']=otp['Bowser Uplift Fuel(LTRS)']*otp['rate']

    avg_fuel_rate_AAZ=otp[otp['Aircraft']=="VT-AAZ"].fuel_upliftment_cost.sum()/otp[otp['Aircraft']=="VT-AAZ"]['Bowser Uplift Fuel(LTRS)'].sum()

    avg_fuel_rate_BAZ=otp[otp['Aircraft']=="VT-BAZ"].fuel_upliftment_cost.sum()/otp[otp['Aircraft']=="VT-BAZ"]['Bowser Uplift Fuel(LTRS)'].sum()

    otp['fuel_cost']=otp.apply(lambda row:avg_fuel_rate_AAZ*row['Total Fuel Burn (Litre)'] if row['Aircraft']=="VT-AAZ" else avg_fuel_rate_BAZ*row['Total Fuel Burn (Litre)'],axis=1)

    cost=cost.merge(otp[['key','fuel_cost']],on='key',how='left')

    cost['fuel_cost_perkg']=cost['fuel_cost']/cost['Total Chargeable Weight']

    qj_mis=qj_mis.merge(cost[['key','fuel_cost_perkg']],on='key',how='left')

    qj_mis['fuel_cost']=qj_mis['Total Chargeable Weight']*qj_mis['fuel_cost_perkg']
    qj_mis.drop(columns=['fuel_cost_perkg'],inplace=True)

    qj_mis.drop(columns=['lane','aircraft','key'],inplace=True)
    #qj_mis.drop(columns=['Airline IataNo','Agent Name','Consignee Name','Shipper - Account name','Agent - Account name','lane','aircraft'],inplace=True)

    qj_mis.rename(columns={'lane_updated':'lane'},inplace=True)

    qj_mis['total_cost']=qj_mis['BH_cost']+qj_mis['fixed_cost']+qj_mis['cycle_cost']+qj_mis['parking_cost']+qj_mis['handling_cost']+qj_mis['cut_cost']+qj_mis['fuel_cost']
    #qj_mis.dropna(inplace=True)

    inputs=pd.DataFrame()
    inputs['cost_component']=''
    inputs['value']=''
    inputs['cost_component']=['fixed_cost','bh_cost','cycle_cost']
    inputs['value']=[fixed_cost,bh_cost,cycle_cost]

    cost=cost.merge(otp[['key','Total Fuel Burn (Litre)']],on='key',how='left')
    cost=cost.merge(otp[['key','Bowser Uplift Fuel(LTRS)']],on='key',how='left')
    cost.drop(columns=['key','block_hours','BH_cost_perkg','fixed_cost_perkg','cycle_cost_perkg','parking_perkg','handling_perkg','fuel_cost_perkg'],inplace=True)
    cost['total']=cost['BH_cost']+cost['fixed_cost']+cost['cycle_cost']+cost['parking']+cost['handling']+cost['cut_cost']+cost['fuel_cost']
    col_order=['Flight Date Local','leg','aircraft','BT','Total Chargeable Weight','fixed_cost','BH_cost','cycle_cost','Bowser Uplift Fuel(LTRS)','Total Fuel Burn (Litre)','fuel_cost','parking','handling','cut_cost','total']
    cost=cost[col_order]
    cost['BT']=cost['BT'].astype(str)
    cost['BT']=cost['BT'].str.slice(0,2).astype(int)+cost['BT'].str.slice(3,5).astype(int)/60



    with pd.ExcelWriter(output_char, engine='xlsxwriter') as writer:
        qj_mis.to_excel(writer, sheet_name='quikjet', index=False)
        inputs.to_excel(writer, sheet_name='inputs', index=False)

        # Access the 'inputs' sheet and write values to specific cells
        workbook = writer.book
        worksheet = writer.sheets['inputs']

        worksheet.write(6, 0, "Parking_charges")
        worksheet.write(14, 0, "Handling_charges")
        worksheet.write(22, 0, "Fuel_charges")
        worksheet.write(29, 0, "cut_ob")
        worksheet.write(36, 0, "cut_ib")

        parking.to_excel(writer,sheet_name='inputs',startrow=7)
        handling.to_excel(writer,sheet_name='inputs',startrow=15)
        fuel_cost.to_excel(writer,sheet_name='inputs',startrow=23)
        cut_ob.to_excel(writer,sheet_name='inputs',startrow=30)
        cut_ib.to_excel(writer,sheet_name='inputs',startrow=37)
        cost.to_excel(writer,sheet_name='finance',index=False)


    print("Running process with the following inputs:")
    print(f"From Date: {a1}/{a2}/{a3}")
    print(f"To Date: {b1}/{b2}/{b3}")
    print(f"Email: {email}")
    output_files = [
        output_file,
        output_char
    ]
    return output_files

